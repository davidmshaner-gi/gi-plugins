"""
Tests for LAND sale comps showing $/Acre in place of $/SF (Excel).

Broker request (Mike Glennon, Lee, 2026-06-02): for LAND sale comps the Excel
should price on $/Acre, not $/SF. $/SF is meaningless for raw land.

Card: davidmshaner-gi/gi-plugins#28.
"""

import os
import tempfile

import pytest
from openpyxl import load_workbook

from conftest import load_skill_helpers

helpers = load_skill_helpers("internal-comps")


def _find_header_row(ws):
    """The header row is the row whose first column is 'Comp ID'."""
    for r in range(1, 8):
        if ws.cell(row=r, column=1).value == "Comp ID":
            return r
    raise AssertionError("header row not found")


def _col_index(ws, header_row, label):
    """1-based column index of the header `label` on `header_row`."""
    c = 1
    while ws.cell(row=header_row, column=c).value is not None:
        if ws.cell(row=header_row, column=c).value == label:
            return c
        c += 1
    raise AssertionError(f"header {label!r} not found")


def _build(rows, asset_type, transaction_type="sale"):
    """Run format_excel and return the loaded Comps worksheet.

    format_excel flattens output_path to a basename and saves into the CWD
    (the Windows long-path guard), returning the actual path it wrote. Run in
    a temp dir and load the returned path.
    """
    validated = {
        "asset_type": asset_type,
        "transaction_type": transaction_type,
        "geography": {"named_market": "Triangle"},
    }
    with tempfile.TemporaryDirectory() as d:
        cwd = os.getcwd()
        os.chdir(d)
        try:
            result = helpers.format_excel(rows, validated, "comps.xlsx", [], [])
            wb = load_workbook(result["path"])
        finally:
            os.chdir(cwd)
    return wb[wb.sheetnames[0]]


# A land sale row with a clean acres value.
LAND_ROW = {
    "comps_id": "L-1",
    "comp_name": "Raw Land Tract",
    "street_address": "123 Farm Rd",
    "city": "Sanford",
    "county": "Lee",
    "acres": 10.0,
    "sale_price": 2_500_000,
    "price_per_sf": 5.74,  # present in data but should be ignored for land
}


def test_land_sale_header_is_price_per_acre():
    ws = _build([LAND_ROW], asset_type="land")
    hr = _find_header_row(ws)
    headers = [ws.cell(row=hr, column=c).value for c in range(1, 24)]
    assert "$/Acre" in headers
    assert "$/SF" not in headers


def test_land_sale_keeps_acres_column():
    ws = _build([LAND_ROW], asset_type="land")
    hr = _find_header_row(ws)
    headers = [ws.cell(row=hr, column=c).value for c in range(1, 24)]
    assert "Acres" in headers  # the Acres column must remain


def test_land_sale_price_per_acre_value_is_computed():
    ws = _build([LAND_ROW], asset_type="land")
    hr = _find_header_row(ws)
    col = _col_index(ws, hr, "$/Acre")
    # 2,500,000 / 10 acres = 250,000
    assert ws.cell(row=hr + 1, column=col).value == pytest.approx(250_000)


def test_land_sale_zero_and_null_acres_render_blank_no_error():
    rows = [
        {**LAND_ROW, "comps_id": "L-zero", "acres": 0},
        {**LAND_ROW, "comps_id": "L-null", "acres": None},
    ]
    ws = _build(rows, asset_type="land")
    hr = _find_header_row(ws)
    col = _col_index(ws, hr, "$/Acre")
    assert ws.cell(row=hr + 1, column=col).value is None  # zero acres -> blank
    assert ws.cell(row=hr + 2, column=col).value is None  # null acres -> blank


def test_land_sale_price_per_acre_number_format_is_currency():
    ws = _build([LAND_ROW], asset_type="land")
    hr = _find_header_row(ws)
    col = _col_index(ws, hr, "$/Acre")
    assert ws.cell(row=hr + 1, column=col).number_format == "$#,##0"


def test_non_land_sale_unchanged_shows_price_per_sf():
    industrial_row = {
        "comps_id": "I-1",
        "comp_name": "Warehouse",
        "street_address": "9 Industry Way",
        "city": "Durham",
        "county": "Durham",
        "square_feet_sold": 50_000,
        "acres": 3.0,
        "sale_price": 6_000_000,
        "price_per_sf": 120.0,
    }
    ws = _build([industrial_row], asset_type="industrial")
    hr = _find_header_row(ws)
    headers = [ws.cell(row=hr, column=c).value for c in range(1, 24)]
    assert "$/SF" in headers
    assert "$/Acre" not in headers
    col = _col_index(ws, hr, "$/SF")
    assert ws.cell(row=hr + 1, column=col).value == pytest.approx(120.0)
