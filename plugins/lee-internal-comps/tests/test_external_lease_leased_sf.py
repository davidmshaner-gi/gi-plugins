"""
External lease comps size on the leased PREMISES, not the building (lee#469).

Christian Sommer (Lee, 2026-08-21) asked for Apex lease comps 2,000–20,000 SF and
got zero rows: the skill sent the broker's size range as `min/max_building_sf`,
which the Worker applied to `building_sf` — the building footprint, NULL on most
external lease rows. The Worker now exposes `leased_sf` (the promoted "Size
Leased SF") and `min/max_leased_sf` on `search_external_lease_comps`; these tests
pin that the skill drives the new params for leases (and still building_sf for
sales) and that every lease size surface — Excel column, Summary stats, chat
table, ranking — reads `leased_sf`.

Card: davidmshaner-gi/lee-and-associates#469 (sibling of Worker 0.43.0).
"""

import os
import tempfile

from openpyxl import load_workbook

from conftest import load_skill_helpers

helpers = load_skill_helpers("external-comps")

LEASE_REQ = {
    "asset_type": "flex", "transaction_type": "lease",
    "geography": {"cities": ["Apex"]},
    "size_range": {"min_sf": 2000, "max_sf": 20000},
    "date_window": {"months": 36},
}
SALE_REQ = {**LEASE_REQ, "transaction_type": "sale", "asset_type": "industrial"}

LEASE_ROW = {
    "external_id": "x1", "property_address": "2019 Production Dr", "property_city": "Apex",
    "county": "Wake", "submarket": "Apex/Holly Springs", "property_type": "Flex",
    "building_sf": None, "leased_sf": 3075, "lease_start_date": "2024-08-01",
    "lease_term_months": 60, "base_rent": 14.5, "rent_type": "NNN", "tenant_name": "Acme",
}
LEASE_ROW_2 = {**LEASE_ROW, "external_id": "x2", "property_address": "1460 Chapel Ridge Rd",
               "building_sf": 47874, "leased_sf": 3240, "base_rent": 15.0}


def test_lease_size_range_maps_to_leased_sf_params():
    v = helpers.validate_request(LEASE_REQ)["validated"]
    out = helpers.build_mcp_params(v)
    assert out["tool_name"] == "search_external_lease_comps"
    p = out["params_list"][0]
    assert p["min_leased_sf"] == 2000 and p["max_leased_sf"] == 20000
    assert "min_building_sf" not in p and "max_building_sf" not in p


def test_sale_size_range_still_maps_to_building_sf_params():
    v = helpers.validate_request(SALE_REQ)["validated"]
    p = helpers.build_mcp_params(v)["params_list"][0]
    assert p["min_building_sf"] == 2000 and p["max_building_sf"] == 20000
    assert "min_leased_sf" not in p


def test_lease_display_columns_carry_leased_sf_before_building_sf():
    cols = helpers.DISPLAY_COLUMNS_LEASE
    assert "leased_sf" in cols
    assert cols.index("leased_sf") < cols.index("building_sf")


def test_lease_ranking_size_proximity_uses_leased_sf():
    v = helpers.validate_request({**LEASE_REQ, "size_range": {"min_sf": 3200, "max_sf": 3300}})["validated"]
    # x2 (3,240) sits nearer the 3,250 midpoint than x1 (3,075); same date, so size decides.
    # A building_sf of 47,874 on x2 must not push it away.
    top, _, _, _ = helpers.rank_comps([LEASE_ROW, LEASE_ROW_2], v)
    assert top[0]["external_id"] == "x2"


def test_lease_markdown_sf_column_is_leased_sf():
    v = helpers.validate_request(LEASE_REQ)["validated"]
    top, uc, sub, undisc = helpers.rank_comps([LEASE_ROW_2], v)
    md = helpers.markdown_table(top, uc, sub, undisc, v)
    assert "3,240" in md
    assert "47,874" not in md


def test_lease_excel_summary_stats_use_leased_sf():
    v = helpers.validate_request(LEASE_REQ)["validated"]
    cwd = os.getcwd()
    with tempfile.TemporaryDirectory() as d:
        os.chdir(d)  # format_excel writes c.xlsx into the CWD (safe_xlsx_name)
        try:
            out = helpers.format_excel([LEASE_ROW, LEASE_ROW_2], v, "out.xlsx", [], [], [])
            wb = load_workbook(out)
        finally:
            os.chdir(cwd)
        summary = wb["Summary"]
        labels = {summary.cell(row=r, column=1).value: summary.cell(row=r, column=2).value
                  for r in range(1, 12)}
        assert labels.get("Avg leased_sf") == (3075 + 3240) / 2
        assert labels.get("Median leased_sf") == (3075 + 3240) / 2
        assert "Avg building_sf" not in labels
