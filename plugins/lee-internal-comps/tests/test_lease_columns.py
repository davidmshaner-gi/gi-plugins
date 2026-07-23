"""
Internal lease comps Excel column layout (gi-plugins#106).

Broker request (Will Fogleman, Lee, 2026-06-17 'Initial Feedback' email,
Internal Comps > Spreadsheet Changes): "Remove Lease Executed."

Removing a display column shifts every positional index after it. The format_excel
number/money/int formatting was historically keyed by hardcoded column POSITION,
which silently mis-formatted columns when the layout changed. These tests pin the
lease header set AND assert that each typed format lands on the correct column by
NAME, so the layout can change without a silent mis-format (the #106 coupling trap).

Card: davidmshaner-gi/gi-plugins#106.
"""

import os
import tempfile

from openpyxl import load_workbook

from conftest import load_skill_helpers

helpers = load_skill_helpers("internal-comps")

LEASE_ROW = {
    "comps_id": "L-1", "comp_name": "Acme Lease", "street_address": "1 Main St",
    "city": "Raleigh", "county": "Wake", "property_type": "Industrial",
    "space_sf": 12500, "building_size": 40000, "square_feet_sold": 40000,
    "lease_execution": "2026-01-15", "lease_commencement": "2026-02-01",
    "term": "60 mo", "lease_type": "NNN", "free_rent_months": 3,
    "ti_allowance_per_sf": 25.0, "asking_rate_per_sf": 12.5, "effective_rate": 11.75,
    "tenant": "BigTenant LLC", "landlord": "LL Co",
    "landlord_rep_agents": "Jane", "tenant_rep_agents": "Bob",
    "link_to_comp_profile": "http://x",
}


def _build_lease(rows):
    validated = {"asset_type": "industrial", "transaction_type": "lease",
                 "geography": {"named_market": "Triangle"}}
    with tempfile.TemporaryDirectory() as d:
        cwd = os.getcwd()
        os.chdir(d)
        try:
            res = helpers.format_excel(rows, validated, "comps.xlsx", [], [])
            wb = load_workbook(res["path"])
        finally:
            os.chdir(cwd)
    return wb[wb.sheetnames[0]]


def _header_row(ws):
    for r in range(1, 8):
        if ws.cell(row=r, column=1).value == "Comp ID":
            return r
    raise AssertionError("header row not found")


def _headers(ws):
    hr = _header_row(ws)
    out = []
    c = 1
    while ws.cell(row=hr, column=c).value is not None:
        out.append(ws.cell(row=hr, column=c).value)
        c += 1
    return out


def _col(ws, label):
    hr = _header_row(ws)
    for h, c in zip(_headers(ws), range(1, 99)):
        if h == label:
            return c
    raise AssertionError(f"header {label!r} not found")


def _fmt(ws, label):
    hr = _header_row(ws)
    return ws.cell(row=hr + 1, column=_col(ws, label)).number_format


# ---------------------------------------------------------------------------
# 1. Lease Executed is gone; the rest of the lease layout is intact + ordered.
# ---------------------------------------------------------------------------

EXPECTED_LEASE_HEADERS = [
    "Comp ID", "Property Type", "Property/Comp", "Address", "City", "County",
    "Leased SF", "Building SF", "Lease Commence", "Term", "Asking $/SF",
    "Effective $/SF", "Lease Type", "Free Rent (mo)", "TI ($/SF)",
    "Tenant", "Landlord", "Landlord Rep", "Tenant Rep", "Comp Profile",
]


def test_lease_executed_column_removed():
    ws = _build_lease([LEASE_ROW])
    assert "Lease Executed" not in _headers(ws)


def test_lease_header_set_and_order():
    ws = _build_lease([LEASE_ROW])
    assert _headers(ws) == EXPECTED_LEASE_HEADERS


# ---------------------------------------------------------------------------
# 2. Typed formats land on the correct columns BY NAME after the shift.
# ---------------------------------------------------------------------------

def test_int_columns_are_thousands_formatted():
    ws = _build_lease([LEASE_ROW])
    for label in ("Leased SF", "Building SF", "Free Rent (mo)"):
        assert _fmt(ws, label) == "#,##0", f"{label} should be int-formatted"


def test_money_per_sf_columns_are_currency_2dp():
    ws = _build_lease([LEASE_ROW])
    for label in ("Asking $/SF", "Effective $/SF", "TI ($/SF)"):
        assert _fmt(ws, label) == "$#,##0.00", f"{label} should be $/SF currency"


def test_text_columns_not_misformatted_as_money():
    """The pre-#106 bug: Lease Type / Tenant / Landlord got money/percent formats
    because the positional index sets were stale. They must stay General."""
    ws = _build_lease([LEASE_ROW])
    for label in ("Lease Type", "Tenant", "Landlord", "Term"):
        assert _fmt(ws, label) == "General", f"{label} should not be numerically formatted"


def test_effective_rate_drives_the_color_scale_column():
    """The color scale must target Effective $/SF, wherever it now sits."""
    ws = _build_lease([LEASE_ROW])
    from openpyxl.utils import get_column_letter
    eff_col_letter = get_column_letter(_col(ws, "Effective $/SF"))
    cf_ranges = [str(rng) for rng in ws.conditional_formatting]
    assert any(eff_col_letter in r for r in cf_ranges), (
        f"color scale should be on Effective $/SF (col {eff_col_letter}); ranges={cf_ranges}"
    )
