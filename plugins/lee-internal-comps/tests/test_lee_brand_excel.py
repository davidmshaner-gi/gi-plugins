"""
Lee-brand the comps Excel deliverables (gi-plugins#90).

A broker opening a generated comps workbook must see Lee & Associates branding on
the FIRST (main comps) sheet: the logo at the top and the official Lee brand color
on the header row -- across BOTH skills (internal-comps + external-comps) and both
transaction types. The official palette is sourced from the same authority as
lee-and-associates#28 (the Brand Guidelines PDF / Worker render repaint):
Lee Red #98002E (PMS 202). Do not invent a second palette.

Before this fix: external-comps used a generic dark-blue header (1F3864) with the
logo only on the Methodology sheet, and internal-comps used a near-but-not-official
maroon (97012D).

Note: this test loads each skill's helpers.py by file path under a unique module
name and does NOT mutate sys.path, so it can't pollute other tests' `import helpers`.
"""

import os
import tempfile

from openpyxl import load_workbook

from conftest import load_skill_helpers

LEE_RED = "98002E"  # official Lee Red, PMS 202 (lee-and-associates#28)

# This file pioneered the unique-module-name load; the loader now lives in
# conftest.py so every test file shares it (gi-plugins#137).
_load_helpers = load_skill_helpers


def _build_main_sheet(skill: str, transaction_type: str):
    """Run the skill's format_excel and return its main (sheet-1) worksheet.

    Both skills flatten the output path to a basename and write into CWD (the
    Windows long-path guard); run in a temp dir and load the path each returns.
    internal-comps returns a dict with 'path'; external-comps returns the path str.
    """
    helpers = _load_helpers(skill)
    validated = {
        "asset_type": "industrial",
        "transaction_type": transaction_type,
        "geography": {"named_market": "Triangle"},
    }
    with tempfile.TemporaryDirectory() as d:
        cwd = os.getcwd()
        os.chdir(d)
        try:
            if skill == "external-comps":
                path = helpers.format_excel([], validated, "comps.xlsx", [], [], [])
            else:
                path = helpers.format_excel([], validated, "comps.xlsx", [], [])["path"]
            wb = load_workbook(path)
        finally:
            os.chdir(cwd)
    return wb[wb.sheetnames[0]]


def _branded_header_cell_present(ws) -> bool:
    """True if any cell in the top band carries a solid fill in official Lee Red."""
    for r in range(1, 9):
        for c in range(1, 26):
            fill = ws.cell(row=r, column=c).fill
            rgb = getattr(getattr(fill, "fgColor", None), "rgb", None)
            if (
                getattr(fill, "patternType", None) == "solid"
                and isinstance(rgb, str)
                and rgb.upper().endswith(LEE_RED)
            ):
                return True
    return False


import pytest


@pytest.mark.parametrize("skill", ["internal-comps", "external-comps"])
@pytest.mark.parametrize("tx", ["sale", "lease"])
def test_main_sheet_has_logo(skill, tx):
    ws = _build_main_sheet(skill, tx)
    assert len(ws._images) >= 1, (
        f"{skill}/{tx}: expected the Lee logo on the main comps sheet, found "
        f"{len(ws._images)} image(s)"
    )


@pytest.mark.parametrize("skill", ["internal-comps", "external-comps"])
@pytest.mark.parametrize("tx", ["sale", "lease"])
def test_main_sheet_header_uses_official_lee_red(skill, tx):
    ws = _build_main_sheet(skill, tx)
    assert _branded_header_cell_present(ws), (
        f"{skill}/{tx}: main-sheet header row is not filled with official "
        f"Lee Red #{LEE_RED}"
    )
