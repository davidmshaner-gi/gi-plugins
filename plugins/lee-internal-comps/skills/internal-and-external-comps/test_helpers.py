from helpers import load_sibling


def test_load_sibling_imports_internal_and_external():
    internal = load_sibling("internal-comps")
    external = load_sibling("external-comps")
    assert hasattr(internal, "validate_request")
    assert hasattr(internal, "build_sql")
    assert hasattr(external, "validate_request")
    assert hasattr(external, "build_mcp_params")


from helpers import to_core, combine, SOURCE_INTERNAL, SOURCE_EXTERNAL


def test_to_core_internal_sale():
    row = {"comps_id": "i1", "street_address": "1 A St", "city": "Garner", "county": "Wake",
           "property_type": "Industrial", "building_size": 50000, "sale_price": 5000000,
           "price_per_sf": 100, "actual_cap_rate": 6.5, "actual_close_date": "2025-03-01"}
    core = to_core(row, SOURCE_INTERNAL, "sale")
    assert core["Source"] == SOURCE_INTERNAL
    assert core["Comp ID"] == "i1"
    assert core["Address"] == "1 A St"
    assert core["Size (SF)"] == 50000
    assert core["Sale Price"] == 5000000
    assert core["$/SF"] == 100
    assert core["Cap Rate"] == 6.5
    assert core["Date"] == "2025-03-01"
    # snake_case stat keys present for summary_stats parity
    assert core["sale_price"] == 5000000 and core["price_per_sf"] == 100
    assert core["square_feet_sold"] == 50000


def test_to_core_external_sale():
    # external_id is a 64-char hash and must NOT be the broker-facing Comp ID;
    # use external_comp_id (short int) / external_property_id instead.
    row = {"external_id": "e31389800c8f344fa0ad7fe718fdf8f498ed52b5a994f6119ed4f50973b32f5c",
           "external_comp_id": 98, "property_address": "2 B St", "property_city": "Cary",
           "county": "Wake", "property_type": "Industrial", "building_sf": 60000,
           "sale_price": 7200000, "price_per_sf": 120, "actual_cap_rate": 6.0,
           "sale_date": "2025-06-01", "external_property_url": "https://ext.example/e1"}
    core = to_core(row, SOURCE_EXTERNAL, "sale")
    assert core["Comp ID"] == 98                       # short id, never the hash
    assert row["external_id"] not in core.values()     # the hash never leaks into any cell
    assert core["City"] == "Cary"
    assert core["Size (SF)"] == 60000
    assert core["Date"] == "2025-06-01"
    assert core["Source URL"] == "https://ext.example/e1"
    assert core["square_feet_sold"] == 60000  # external building_sf mapped to stat key


def test_to_core_external_comp_id_prefers_source_property_id_then_falls_back():
    # external_property_id wins when present
    r1 = {"external_id": "hash", "external_comp_id": 7, "external_property_id": "CS-12345"}
    assert to_core(r1, SOURCE_EXTERNAL, "sale")["Comp ID"] == "CS-12345"
    # falls back to external_comp_id when no source property id
    r2 = {"external_id": "hash", "external_comp_id": 7}
    assert to_core(r2, SOURCE_EXTERNAL, "sale")["Comp ID"] == 7
    # never the hash
    r3 = {"external_id": "e31389800c8f344fa0ad7fe718fdf8f498ed52b5a994f6119ed4f50973b32f5c"}
    assert to_core(r3, SOURCE_EXTERNAL, "sale")["Comp ID"] == ""


def test_to_core_internal_lease():
    row = {"comps_id": "i2", "street_address": "3 C St", "city": "Apex", "county": "Wake",
           "property_type": "Industrial", "space_sf": 12000, "effective_rate": 9.5,
           "asking_rate_per_sf": 10.0, "lease_execution": "2025-04-01", "lease_type": "NNN"}
    core = to_core(row, SOURCE_INTERNAL, "lease")
    assert core["Leased SF"] == 12000
    assert core["Rent"] == 9.5
    assert core["Date"] == "2025-04-01"
    assert core["Lease Type"] == "NNN"
    # The old W1 building-size-substitution flag is gone (gi-plugins#105).
    assert "_leased_sf_is_building_size" not in core
    assert core["effective_rate"] == 9.5 and core["space_sf"] == 12000


def test_to_core_external_lease_leased_sf_is_blank():
    # gi-plugins#105: The external platform carries no true leased area for leases, so "Leased SF"
    # must render BLANK for external lease rows — never the building size.
    row = {"external_id": "e2", "property_address": "4 D St", "property_city": "Apex",
           "county": "Wake", "property_type": "Industrial", "building_sf": 40000,
           "base_rent": 11.0, "lease_start_date": "2025-05-01", "rent_type": "Gross"}
    core = to_core(row, SOURCE_EXTERNAL, "lease")
    assert core["Leased SF"] is None
    assert core["Leased SF"] != 40000          # building size must not leak in
    assert core["Rent"] == 11.0
    assert core["Lease Type"] == "Gross"
    assert "_leased_sf_is_building_size" not in core
    # space_sf stat key mirrors the (now-blank) leased value.
    assert core["effective_rate"] == 11.0 and core["space_sf"] is None


def test_combine_keeps_both_rows_and_sorts_desc():
    internal = [to_core({"comps_id": "i1", "street_address": "1 A St",
                         "actual_close_date": "2025-01-01", "building_size": 1,
                         "sale_price": 1, "price_per_sf": 1}, SOURCE_INTERNAL, "sale")]
    external = [to_core({"external_id": "e1", "property_address": "1 A St",
                         "sale_date": "2025-09-01", "building_sf": 1,
                         "sale_price": 1, "price_per_sf": 1}, SOURCE_EXTERNAL, "sale")]
    out = combine(internal, external, "sale")
    assert len(out) == 2                       # same address in both → BOTH kept (no dedup)
    assert out[0]["Date"] == "2025-09-01"      # most recent first
    assert {r["Source"] for r in out} == {SOURCE_INTERNAL, SOURCE_EXTERNAL}


from helpers import unified_markdown_table, format_unified_excel


def _sale_combo():
    internal = [to_core({"comps_id": "i1", "street_address": "1 A St",
                         "actual_close_date": "2025-01-01", "building_size": 50000,
                         "sale_price": 5000000, "price_per_sf": 100, "property_type": "Industrial",
                         "city": "Garner", "county": "Wake"}, SOURCE_INTERNAL, "sale")]
    external = [to_core({"external_id": "e1", "property_address": "2 B St",
                         "sale_date": "2025-02-01", "building_sf": 60000,
                         "sale_price": 7200000, "price_per_sf": 120, "property_type": "Industrial",
                         "property_city": "Cary", "county": "Wake"}, SOURCE_EXTERNAL, "sale")]
    return combine(internal, external, "sale"), internal, external


def test_unified_markdown_table_has_source_column():
    rows, _, _ = _sale_combo()
    md = unified_markdown_table(rows, {"comp_type": "sale"})
    assert md.splitlines()[0].startswith("| Source")
    assert "Internal — Dealius" in md
    assert "External" in md


def test_unified_markdown_external_lease_blanks_leased_sf_no_footnote():
    # gi-plugins#105: external lease Leased SF is blank (the external platform has no leased area),
    # so the building-size value never appears and the old W1 footnote is gone.
    ext = [to_core({"external_id": "e2", "property_address": "4 D St", "building_sf": 40000,
                    "base_rent": 11.0, "lease_start_date": "2025-05-01", "rent_type": "Gross"},
                   SOURCE_EXTERNAL, "lease")]
    md = unified_markdown_table(combine([], ext, "lease"), {"comp_type": "lease"})
    assert "Leased SF" in md.splitlines()[0]
    assert "40000" not in md                       # building size must not leak in
    assert "building size, not leased area" not in md


def test_format_unified_excel_writes_three_sheets(tmp_path):
    import os
    core, internal, external = _sale_combo()
    # The Windows long-path guard (gi-plugins#7) flattens the output to a basename
    # in the CWD and returns the actual path written; run in tmp_path and load that.
    cwd = os.getcwd()
    os.chdir(tmp_path)
    try:
        written = format_unified_excel(
            core, internal_native=[{"comps_id": "i1"}],
            external_native=[{"external_id": "e1"}],
            validated={"comp_type": "sale"},
            xlsx_path="comps-all-industrial-2026-06-02.xlsx")
        assert os.sep not in written  # flattened to a CWD basename
        import openpyxl
        wb = openpyxl.load_workbook(written)
    finally:
        os.chdir(cwd)
    assert set(wb.sheetnames) >= {"All Comps", "Internal (Dealius)", "External"}
    assert wb["All Comps"].cell(row=1, column=1).value == "Source"


# ---------------------------------------------------------------------------
# gi-plugins#62 — combine() must sort by a PARSED date, not the raw string.
# Internal dates are MM/DD/YYYY (actual_close_date / lease_execution); external
# dates are ISO YYYY-MM-DD (sale_date / lease_start_date). A raw string sort
# interleaves the two formats wrong (e.g. "12/01/2025" sorts after "2026-..").
# ---------------------------------------------------------------------------

def test_combine_sorts_mixed_date_formats_chronologically():
    # Internal rows carry MM/DD/YYYY; external rows carry ISO YYYY-MM-DD.
    internal = [
        to_core({"comps_id": "INT-NEWEST", "actual_close_date": "06/01/2026",
                 "sale_price": 1, "price_per_sf": 1, "building_size": 1},
                SOURCE_INTERNAL, "sale"),
        to_core({"comps_id": "INT-OLD", "actual_close_date": "12/01/2025",
                 "sale_price": 1, "price_per_sf": 1, "building_size": 1},
                SOURCE_INTERNAL, "sale"),
    ]
    external = [
        to_core({"external_property_id": "EXT-MID", "sale_date": "2026-01-15",
                 "sale_price": 1, "price_per_sf": 1, "building_sf": 1},
                SOURCE_EXTERNAL, "sale"),
        to_core({"external_property_id": "EXT-OLDEST", "sale_date": "2024-03-10",
                 "sale_price": 1, "price_per_sf": 1, "building_sf": 1},
                SOURCE_EXTERNAL, "sale"),
    ]
    out = combine(internal, external, "sale")
    order = [r["Comp ID"] for r in out]
    # True chronological, most-recent first — NOT what a raw string sort produces.
    assert order == ["INT-NEWEST", "EXT-MID", "INT-OLD", "EXT-OLDEST"]


def test_combine_blank_or_unparseable_dates_sort_last():
    internal = [
        to_core({"comps_id": "HAS-DATE", "actual_close_date": "03/15/2026",
                 "sale_price": 1, "price_per_sf": 1, "building_size": 1},
                SOURCE_INTERNAL, "sale"),
        to_core({"comps_id": "NO-DATE", "actual_close_date": "",
                 "sale_price": 1, "price_per_sf": 1, "building_size": 1},
                SOURCE_INTERNAL, "sale"),
    ]
    external = [
        to_core({"external_property_id": "BAD-DATE", "sale_date": "not-a-date",
                 "sale_price": 1, "price_per_sf": 1, "building_sf": 1},
                SOURCE_EXTERNAL, "sale"),
    ]
    out = combine(internal, external, "sale")
    assert len(out) == 3                       # nothing dropped
    assert out[0]["Comp ID"] == "HAS-DATE"     # the only real date comes first
    # The two blank/unparseable rows sort to the bottom (order among them unspecified).
    assert {out[1]["Comp ID"], out[2]["Comp ID"]} == {"NO-DATE", "BAD-DATE"}
