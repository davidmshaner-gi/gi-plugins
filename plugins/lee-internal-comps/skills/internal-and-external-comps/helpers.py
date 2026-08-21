"""Helpers for the internal-and-external-comps skill — the unified "all comps" default.

This skill is a thin orchestrator: it reuses the parse/validate/query logic of the
sibling `internal-comps` and `external-comps` skills (via load_sibling) and adds only
the combine layer (to_core, combine, format_unified_excel, unified_markdown_table).
"""
import importlib.util
import sys
from datetime import date, datetime
from pathlib import Path


def load_sibling(skill_name: str):
    """Import a sibling skill's helpers.py by skill-dir name.

    Sibling skills live at ../<skill_name>/helpers.py relative to this file. Returns the
    imported module. Cached in sys.modules so repeated calls are cheap and idempotent.
    """
    path = Path(__file__).resolve().parent.parent / skill_name / "helpers.py"
    mod_name = f"_sibling_{skill_name.replace('-', '_')}"
    if mod_name in sys.modules:
        return sys.modules[mod_name]
    spec = importlib.util.spec_from_file_location(mod_name, path)
    if spec is None or spec.loader is None:
        raise ImportError(f"load_sibling: cannot build import spec for {path}")
    mod = importlib.util.module_from_spec(spec)
    sys.modules[mod_name] = mod
    spec.loader.exec_module(mod)
    return mod


# =====================================================================
# Core schema + combine  (the unified "all comps" merge layer)
# =====================================================================

SOURCE_INTERNAL = "Internal — Dealius"
SOURCE_EXTERNAL = "External"


def _first(row: dict, *keys):
    """First non-empty value among keys (None/"" treated as empty)."""
    for k in keys:
        v = row.get(k)
        if v not in (None, ""):
            return v
    return None


def to_core(row: dict, source: str, tx_type: str) -> dict:
    """Map one source row (internal Dealius or external) to the unified CORE schema.

    The core row carries BOTH display-label keys (what the unified table/Excel/PDF render)
    AND the snake_case stat keys that the MCP `computeSummaryStats` reads server-side
    (sale: sale_price/price_per_sf/square_feet_sold; lease: effective_rate/
    asking_rate_per_sf/space_sf). Internal rows already use those snake_case names; external
    rows are mapped onto them here so stats work regardless of source.
    """
    is_internal = source.startswith("Internal")
    if is_internal:
        comp_id = row.get("comps_id")
        address = row.get("street_address")
        city = row.get("city")
        url = ""
    else:
        # Broker-readable id: prefer the external platform's id, then the short external_comp_id.
        # NEVER external_id (a 64-char address hash that overflows the comp table).
        comp_id = row.get("external_property_id") or row.get("external_comp_id") or ""
        address = row.get("property_address")
        city = row.get("property_city")
        url = row.get("external_property_url") or ""
    county = row.get("county")
    asset = row.get("property_type")

    if tx_type == "sale":
        if is_internal:
            size = _first(row, "square_feet_sold", "building_size")
            date = row.get("actual_close_date")
        else:
            size = row.get("building_sf")
            date = row.get("sale_date")
        cap = _first(row, "actual_cap_rate", "asking_cap_rate")
        sale_price = row.get("sale_price")
        ppsf = row.get("price_per_sf")
        return {
            "Source": source,
            "Comp ID": comp_id,
            "Address": address,
            "City": city,
            "County": county,
            "Asset Type": asset,
            "Size (SF)": size,
            "Sale Price": sale_price,
            "$/SF": ppsf,
            "Cap Rate": cap,
            "Date": date,
            "Source URL": url,
            # snake_case stat keys (server-side computeSummaryStats parity)
            "sale_price": sale_price,
            "price_per_sf": ppsf,
            "square_feet_sold": size,
        }

    # lease
    if is_internal:
        leased = _first(row, "space_sf", "square_feet_sold")
        rent = _first(row, "effective_rate", "asking_rate_per_sf")
        asking = row.get("asking_rate_per_sf")
        date = _first(row, "lease_execution", "lease_commencement")
        lease_type = row.get("lease_type")
    else:
        # leased_sf is the external platform's "Size Leased SF", promoted to a
        # typed column and backfilled (lee#469, closes lee#180). building_sf is the
        # building footprint and must NEVER stand in for the leased premises
        # (gi-plugins#105, Will Fogleman 2026-06-17): a row without leased_sf
        # renders BLANK rather than mislabel building size. 0 is an unknown-value
        # placeholder, never a real size (gi-plugins#82), so it renders blank too.
        leased = row.get("leased_sf") or None
        rent = row.get("base_rent")
        asking = None
        date = row.get("lease_start_date")
        lease_type = row.get("rent_type")
    return {
        "Source": source,
        "Comp ID": comp_id,
        "Address": address,
        "City": city,
        "County": county,
        "Asset Type": asset,
        "Leased SF": leased,
        "Rent": rent,
        "Date": date,
        "Lease Type": lease_type,
        "Source URL": url,
        # snake_case stat keys (server-side computeSummaryStats parity)
        "effective_rate": rent,
        "asking_rate_per_sf": asking,
        "space_sf": leased,
    }


def _parse_date(value) -> "date | None":
    """Parse a core-row Date into a real date for chronological sorting (gi-plugins#62).

    Internal rows carry MM/DD/YYYY (from actual_close_date / lease_execution); external
    rows carry ISO YYYY-MM-DD (from sale_date / lease_start_date). A naive string sort
    interleaves the two formats wrong (e.g. "12/01/2020" sorts after "2026-01-01").
    Returns None for missing/blank/unparseable values so they sort LAST, never crash.
    """
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    # Trim a time component if one rode along (e.g. "2026-01-15 00:00:00").
    s = s.split("T")[0].split(" ")[0]
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def combine(internal_core: list, external_core: list, tx_type: str) -> list:
    """Concatenate internal + external core rows, sort most-recent-Date first. NO dedup —
    a property present in both sources stays as two rows, each tagged by Source.

    Dates are parsed to real dates before sorting (gi-plugins#62) because internal
    (MM/DD/YYYY) and external (ISO) formats can't be compared as raw strings. Rows with
    a missing/blank/unparseable Date sort last.
    """
    rows = list(internal_core) + list(external_core)

    def _key(r):
        # (has_date, date) so real dates sort most-recent-first and blanks fall to the
        # bottom under reverse=True (False < True, and date.min anchors the blanks).
        d = _parse_date(r.get("Date"))
        return (d is not None, d or date.min)

    rows.sort(key=_key, reverse=True)
    return rows


# =====================================================================
# Unified output: chat table + Excel
# =====================================================================

CORE_COLUMNS_SALE = ["Source", "Comp ID", "Address", "City", "County", "Asset Type",
                     "Size (SF)", "Sale Price", "$/SF", "Cap Rate", "Date"]
CORE_COLUMNS_LEASE = ["Source", "Comp ID", "Address", "City", "County", "Asset Type",
                      "Leased SF", "Rent", "Date", "Lease Type"]

LEE_BRAND_MAROON = "97012D"


def _core_columns(tx_type: str) -> list:
    return CORE_COLUMNS_SALE if tx_type == "sale" else CORE_COLUMNS_LEASE


def _fmt(value) -> str:
    """Plain display formatting for a cell value (markdown + Excel fallback)."""
    if value is None or value == "":
        return ""
    return str(value)


def unified_markdown_table(core_rows: list, validated: dict) -> str:
    """Combined chat table over the core columns, Source first.

    External lease rows show the leased premises (`leased_sf`, lee#469) or blank when
    it is absent — never the building size (gi-plugins#105), so there is no
    building-size-substitution footnote.
    """
    tx_type = validated.get("comp_type") or validated.get("transaction_type") or "sale"
    cols = _core_columns(tx_type)
    lines = ["| " + " | ".join(cols) + " |",
             "| " + " | ".join("---" for _ in cols) + " |"]
    for r in core_rows:
        lines.append("| " + " | ".join(_fmt(r.get(c)) for c in cols) + " |")
    md = "\n".join(lines)
    return md


def format_unified_excel(core_rows: list, internal_native: list, external_native: list,
                         validated: dict, xlsx_path: str) -> str:
    """Write a 3-sheet workbook: 'All Comps' (core rows, Source first) + native detail
    sheets for each source. Returns xlsx_path. Lee-maroon header row on every sheet."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    header_fill = PatternFill("solid", start_color=LEE_BRAND_MAROON)
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
    tx_type = validated.get("comp_type") or validated.get("transaction_type") or "sale"

    def _write_sheet(ws, columns, rows):
        for ci, col in enumerate(columns, start=1):
            c = ws.cell(row=1, column=ci, value=col)
            c.fill = header_fill
            c.font = header_font
            c.alignment = wrap
        for ri, row in enumerate(rows, start=2):
            for ci, col in enumerate(columns, start=1):
                ws.cell(row=ri, column=ci, value=row.get(col))

    wb = Workbook()
    # Sheet 1 — All Comps (core)
    ws_all = wb.active
    ws_all.title = "All Comps"
    _write_sheet(ws_all, _core_columns(tx_type), core_rows)

    # Sheet 2 — Internal (Dealius) native
    ws_int = wb.create_sheet("Internal (Dealius)")
    int_cols = list(internal_native[0].keys()) if internal_native else ["(no internal rows)"]
    _write_sheet(ws_int, int_cols, internal_native)

    # Sheet 3 — External native
    ws_ext = wb.create_sheet("External")
    ext_cols = list(external_native[0].keys()) if external_native else ["(no external rows)"]
    _write_sheet(ws_ext, ext_cols, external_native)

    # Windows 218-char path guard (gi-plugins#7). This orchestrator already loads
    # its siblings, so reuse the canonical helper rather than duplicating it:
    # flatten to a CWD basename + cap the filename so a deep/long path can't survive.
    xlsx_path = load_sibling("internal-comps").safe_xlsx_name(xlsx_path)
    wb.save(xlsx_path)
    return xlsx_path
