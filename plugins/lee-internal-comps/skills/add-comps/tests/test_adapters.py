"""
Tests for the add-comps skill adapters + payload builders (B4-B6).

B4 — spreadsheet (.xlsx/.csv) + pasted-text adapters.
B5 — image + LLM-fallback adapters (dependency-injected model-vision callable).
B6 — dry-run summary + the lee_comps_add_write payload builder.

Pure-Python; the model orchestrates MCP, the helpers are deterministic.
Run from the skill dir: python3 -m pytest tests/ -v
"""

import os
import sys

import pytest

# Import helpers.py from the skill dir (parent of tests/).
sys.path.insert(
    0, os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
)

from helpers import (  # noqa: E402
    PARSER_VERSION,
    AmbiguousSheetError,
    build_write_payload,
    dry_run_summary,
    llm_fallback,
    parse_email,
    parse_image,
    parse_spreadsheet,
    parse_text,
)

FIXTURES = os.path.join(os.path.dirname(__file__), "fixtures")


# ---------------------------------------------------------------------------
# B4 — spreadsheet + text adapters
# ---------------------------------------------------------------------------

def test_parse_spreadsheet_csv_routes_to_sale_block():
    """The synthetic sale CSV routes to the sale block: sale_price set,
    base_rent absent/None, and every row validated."""
    rows = parse_spreadsheet(os.path.join(FIXTURES, "sale_comps.csv"))
    assert len(rows) == 3
    assert all(r["transaction_type"] == "sale" for r in rows)
    assert all(r.get("sale_price") for r in rows)
    # Lease block must be absent/None on a sale row.
    assert all(r.get("base_rent") is None for r in rows)
    # First row coerced numerics.
    first = rows[0]
    assert first["sale_price"] == 6_000_000
    assert first["sale_price_per_sf"] == 120.0
    assert first["buyer"] == "Acme REIT"
    assert first["property_address"] == "101 Synthetic Way"
    # Validated.
    assert all("flagged" in r for r in rows)
    assert all(r["flagged"] in (0, 1) for r in rows)


def test_parse_spreadsheet_xlsx_single_sale_sheet(tmp_path):
    """A single-sheet sale xlsx routes to the sale block."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales"
    ws.append(["Address (Park)", "City", "Sale Price", "$/SF",
               "Sale Date", "Cap Rate", "Buyer", "Seller"])
    ws.append(["1 Foo St", "Durham", "$6,000,000", "$120.00",
               "2025-03-01", "6.5%", "Acme REIT", "Old Owner LLC"])
    path = tmp_path / "one_sheet.xlsx"
    wb.save(path)

    rows = parse_spreadsheet(str(path))
    assert len(rows) == 1
    assert rows[0]["transaction_type"] == "sale"
    assert rows[0]["sale_price"] == 6_000_000
    assert rows[0].get("base_rent") is None


def test_parse_spreadsheet_skips_prose_tabs(tmp_path):
    """A multi-tab xlsx with one comp sheet + a prose/view tab picks only
    the comp-shaped sheet (no ambiguity)."""
    import openpyxl
    wb = openpyxl.Workbook()
    comp = wb.active
    comp.title = "Comps"
    comp.append(["Address (Park)", "City", "Sale Price", "Sale Date",
                 "Cap Rate", "Buyer", "Seller"])
    comp.append(["1 Foo St", "Durham", "$6,000,000", "2025-03-01",
                 "6.5%", "Acme REIT", "Old Owner LLC"])
    notes = wb.create_sheet("Read Me")
    notes.append(["These comps were pulled from CoStar in May."])
    notes.append(["Contact the analyst with questions."])
    path = tmp_path / "with_prose.xlsx"
    wb.save(path)

    rows = parse_spreadsheet(str(path))
    assert len(rows) == 1
    assert rows[0]["transaction_type"] == "sale"


def test_parse_spreadsheet_ambiguous_multi_sheet_raises(tmp_path):
    """Two comp-shaped sheets with no obvious single pick raise
    AmbiguousSheetError carrying the candidate sheet names."""
    import openpyxl
    wb = openpyxl.Workbook()
    s1 = wb.active
    s1.title = "Sale Comps"
    s1.append(["Address (Park)", "City", "Sale Price", "Sale Date",
               "Cap Rate", "Buyer", "Seller"])
    s1.append(["1 Foo St", "Durham", "$6,000,000", "2025-03-01",
               "6.5%", "Acme REIT", "Old Owner LLC"])
    s2 = wb.create_sheet("Lease Comps")
    s2.append(["Address (Park)", "City", "Tenant", "Sign Date",
               "SF", "Year One Base Rent NNN", "Term (months)"])
    s2.append(["2 Bar St", "Raleigh", "Tenant A", "2025Q2",
               "25,500", "$13.07", "60"])
    path = tmp_path / "ambiguous.xlsx"
    wb.save(path)

    with pytest.raises(AmbiguousSheetError) as exc:
        parse_spreadsheet(str(path))
    assert "Sale Comps" in exc.value.sheet_names
    assert "Lease Comps" in exc.value.sheet_names


def test_parse_text_pipe_table():
    """A pasted pipe-delimited table parses through detect+alias+validate."""
    blob = (
        "Address (Park) | City | Sale Price | Sale Date | Cap Rate | Buyer | Seller\n"
        "1 Foo St | Durham | $6,000,000 | 2025-03-01 | 6.5% | Acme REIT | Old Owner LLC\n"
        "2 Bar St | Raleigh | $3,250,000 | 2025-01-22 | 7.1% | Harbor Industrial | Sunset Partners\n"
    )
    rows = parse_text(blob)
    assert len(rows) == 2
    assert all(r["transaction_type"] == "sale" for r in rows)
    assert rows[0]["sale_price"] == 6_000_000
    assert rows[0]["buyer"] == "Acme REIT"
    assert all("flagged" in r for r in rows)


def test_parse_text_tab_table_lease():
    """A pasted tab-delimited lease table parses and routes to lease."""
    blob = (
        "Address (Park)\tCity\tTenant\tSign Date\tSF\tYear One Base Rent NNN\tTerm (months)\n"
        "9 Lease Ln\tCary\tTenant Z\t2025Q3\t25,500\t$13.07\t60\n"
    )
    rows = parse_text(blob)
    assert len(rows) == 1
    assert rows[0]["transaction_type"] == "lease"
    assert rows[0]["leased_sf"] == 25500
    assert rows[0]["base_rent"] == 13.07


# ---------------------------------------------------------------------------
# B5 — image + LLM-fallback adapters (injected model-vision callable)
# ---------------------------------------------------------------------------

def _fake_model_two_canonical_rows(_ref):
    """A mock model_extract: returns two already-canonical sale rows."""
    return [
        {
            "transaction_type": "sale",
            "property_address": "1 Vision St",
            "sale_price": 6_000_000,
            "sale_date": "2025-03-01",
            "buyer": "Acme REIT",
        },
        {
            "transaction_type": "sale",
            "property_address": "2 Vision St",
            "sale_price": 3_250_000,
            "sale_date": "2025-01-22",
            "buyer": "Harbor Industrial",
        },
    ]


def test_parse_image_validates_mocked_rows():
    rows = parse_image("fake://image.png", _fake_model_two_canonical_rows)
    assert len(rows) == 2
    assert all(r["transaction_type"] == "sale" for r in rows)
    # Ran through validate_row.
    assert all("flagged" in r for r in rows)
    assert all(r["flagged"] in (0, 1) for r in rows)
    # A complete sale row is not flagged.
    assert rows[0]["flagged"] == 0


def test_llm_fallback_validates_mocked_rows():
    def fake_model(_blob):
        return _fake_model_two_canonical_rows(_blob)

    rows = llm_fallback("some unparseable blob", fake_model)
    assert len(rows) == 2
    assert all("flagged" in r for r in rows)


def test_parse_image_missing_model_raises():
    """With no injected model and no default, the adapter raises clearly."""
    with pytest.raises(ValueError):
        parse_image("fake://image.png", None)


# ---------------------------------------------------------------------------
# B6 — dry-run summary + write-payload builder
# ---------------------------------------------------------------------------

def _golden_rows():
    html = open(os.path.join(FIXTURES, "silas_email.html")).read()
    return parse_email(html)


def test_dry_run_summary_on_golden_rows():
    rows = _golden_rows()
    summary = dry_run_summary(rows)
    assert summary["total"] == 30
    assert summary["by_source"] == {
        "JLL": 12, "Foundry": 4, "Tri Property": 9, "Prologis": 5,
    }
    # flagged is a list of {row index/reason} entries; structure check only.
    assert isinstance(summary["flagged"], list)
    for f in summary["flagged"]:
        assert "flag_reason" in f


def test_dry_run_summary_reports_flagged():
    rows = [
        {"transaction_type": "sale", "original_source": "X",
         "flagged": 1, "flag_reason": "missing required fields: sale_price"},
        {"transaction_type": "sale", "original_source": "X",
         "flagged": 0, "flag_reason": ""},
    ]
    summary = dry_run_summary(rows)
    assert summary["total"] == 2
    assert summary["by_source"] == {"X": 2}
    assert len(summary["flagged"]) == 1
    assert "sale_price" in summary["flagged"][0]["flag_reason"]


def test_build_write_payload_shape():
    rows = _golden_rows()
    meta = {
        "added_by": "moss@lee.com",
        "import_method": "email_paste",
        "source_label": "Silas full 2025 industrial set",
        "raw_blob": "<html>...</html>",
        "client_id": "lee",
        "notes": "forwarded by Moss",
    }
    payload = build_write_payload(rows, meta)
    # Required keys present.
    for key in ("added_by", "import_method", "raw_blob", "parser_version", "rows"):
        assert key in payload
    assert payload["added_by"] == "moss@lee.com"
    assert payload["import_method"] == "email_paste"
    assert payload["parser_version"] == PARSER_VERSION
    assert payload["rows"] is rows or payload["rows"] == rows
    assert len(payload["rows"]) == 30
    # Optional keys passed through.
    assert payload["client_id"] == "lee"
    assert payload["source_label"] == "Silas full 2025 industrial set"
    assert payload["notes"] == "forwarded by Moss"


def test_build_write_payload_omits_absent_optionals():
    rows = []
    meta = {
        "added_by": "moss@lee.com",
        "import_method": "text_paste",
        "raw_blob": "1 Foo | ...",
    }
    payload = build_write_payload(rows, meta)
    assert payload["added_by"] == "moss@lee.com"
    assert payload["import_method"] == "text_paste"
    assert "client_id" not in payload
    assert "source_label" not in payload
    assert "notes" not in payload
    assert "raw_blob_ref" not in payload


def test_build_write_payload_rejects_bad_import_method():
    with pytest.raises(ValueError):
        build_write_payload([], {"added_by": "x", "import_method": "bogus",
                                 "raw_blob": ""})
