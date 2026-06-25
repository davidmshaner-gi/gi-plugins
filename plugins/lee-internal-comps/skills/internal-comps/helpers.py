"""
helpers.py — atomic helpers for the internal-comps skill.

Run in the Cowork sandbox. The model orchestrates; helpers are deterministic.
None of these helpers call MCP tools — SQL execution and email sending are
the model's responsibility (the model has MCP access; the sandbox does not).

Design contract:
  - Open-shaped dicts in / dicts out. Helpers tolerate extra keys.
  - Three load-bearing keys on the request: asset_type, transaction_type,
    geography. Everything else is optional and may be absent.
  - Frozen output layout — see SKILL.md "Output" section. Do not parameterize
    beyond what these signatures expose.
"""

import base64
import os
import tempfile
from datetime import date
from typing import Optional, Literal

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.drawing.image import Image as XLImage


# The deliverable lands in Cowork's per-session output directory, which on Windows
# already runs ~200 chars deep. Excel refuses to OPEN any workbook whose full path
# exceeds 218 chars ("the file path is too long") — stricter than Windows' own 260.
# That session dir's length is fixed by Cowork and the file MUST land inside it (it
# is where the broker sees the deliverable), so the ONLY lever the skill controls is
# the filename. Bonner's 2026-06-25 Cowork measurement put his Windows session output
# dir near the ceiling — his pasted path measured 188 but he reported ~210, leaving
# only ~8-30 chars under 218; we size for the worst case. So even `comps.xlsx` (full
# path ~199-221 depending on that base) is too risky. We IGNORE any descriptive name
# the caller builds and emit the shortest practical stub, `c.xlsx` (6 chars; `c1.xlsx`,
# ..., `c99.xlsx` stay ≤ 8), which clears 218 for any base ≤ 211. The descriptive title
# still rides on the Sheet 1 tab name; the broker renames the file to taste (Cowork can
# do it on request).
XLSX_STUB = "c"  # stub + enum suffix + ".xlsx" must stay <=~8 chars (the budget on the deepest session dirs); "c" leaves room through c99.xlsx


def safe_xlsx_name(path: str = "") -> str:
    """Return the shortest stable .xlsx filename in the CWD, enumerating on collision.

    Emits `c.xlsx` (XLSX_STUB + ".xlsx"); if that name already exists in the CWD (a
    second comps pull in the same Cowork session), enumerates `c1.xlsx`, `c2.xlsx`, ...
    so a later pull never silently clobbers an earlier deliverable. `path` is accepted for
    back-compat with callers that still pass a descriptive path, and deliberately
    ignored — see the XLSX_STUB note above for why the descriptive name can't survive
    the Windows 218-char Excel-open limit (gi-plugins#7). Canonical here; the
    external-comps skill keeps an identical mirror (it takes no sibling import).
    """
    candidate = f"{XLSX_STUB}.xlsx"
    n = 1
    while os.path.exists(candidate):
        candidate = f"{XLSX_STUB}{n}.xlsx"
        n += 1
    return candidate


LEE_BRAND_MAROON = "98002E"  # official Lee Red, PMS 202 (lee-and-associates#28 / Brand Guidelines)
LEE_LOGO_FILENAME = "lee_logo.png"  # ships alongside helpers.py (used in local dev)

# Lee & Associates logo, base64-encoded. Source: lee_logo.png in this skill bundle.
# Embedded here (rather than fetched at runtime) to eliminate the bundle-distribution
# reliability bug AND avoid any outbound-HTTPS dependency (Cowork sandbox has no egress).
# Update via scripts/encode_logo.py if logo changes.
LEE_LOGO_B64 = """\
iVBORw0KGgoAAAANSUhEUgAAAPAAAABJCAYAAAADkKsvAABCbklEQVR42u29eZxcRbk+/rxV55xe
ZkkYQkICsiiLzkAAAXcICi6g6PeqPaJeBVQSQQJiwJBkuqureyYhhIgRERLZXK7LtKLg9YoKaFhU
RC4IZNhUVhNIwkwmM9PLOafq/f3Rpyc9k8kyIUH83XnzmU+Snj7n1Kmqt979eQkA7p96wkeSJHMV
tq0McrB9YgJCF+KvgzAL377+rt8yIAiwmKAJmqBXlcQD0048PQ6nANDMnWBeACAGXEE4rhHylj/v
c8I7CLDdSMmJ6ZygCXqVGZgZVwiCW2JjxnNhiW0QI5FgIAsAKbTyxHRO0AS9ygwM4JAyWybQeCWo
LMIwA4f+Bce6BG0ZoIkpnaAJehUZmEACu8B4DIYFyAJi4MDGCfV5giboXySBMX7mrXqsGIwJvXmC
JuhfR86uXGQj1rUgmFeBhauquXpF6nkWevjf+tX3mBNzdZ6IaKcmjJl36n3HuN/ws3bh2p26rBsp
sQbrCQDasJrbATPBSv8mDGzrJDDtYQncjZRMoZUJ2gJ6dz6KFECvIiMz0fjOn11krl161s6SAoQG
bDsKZlu/m2Cp1zAD87DtW/0RoD3GwN1IydpGeXT/97X4YXFfa4V06ja2YSMEyzHNAL8qimgGxfhl
qpiHw7J5kcsidLm4YP0j/9AAM0CEPa9CLF26tKlYpAbPC9j3/T6ttb+9789eudI9cOPGlsEd3LcR
QN39CADPv+yySfGSSPhewDu4LtBa945HC4pi/bQKR76LgcMJxAKy5wt48I+6dqZjwqp6zTJwlXFr
jFxdMQsLgtzdKrMgFMwDU098uwfxlcD3T2KIvQASYd32IEgaS9FkAB4IIAr/YjeXX+CKMySMG4AJ
FpX8lCP+UKbyHNrwt7/vScmhlHK01uFg0U/H47EvhiFxCOeDAO5JpbplodA+QpKlUilZKBTMlLVr
jzVu/NdOGFowE4jMWDq29TyC43wMwO+VUq7W2o8VK1/zEomUrdiAwMQYJcmj65icvwJ4z85KXgLs
N3HU0XFgFYGOb4jWfAgG1+Gouw3Cs+dgzd8nJPFrlIFtpC7XM3HtM9qtzJuShIK5f+oJHR7JnAdB
hg1oG08Z67in6k/4Vx4oPo9yo4AQNYebASfiQp5cMd7/qJa3HI/ePw/ueUnMSSllk7UWgNnhnBPg
CCGaCYDreXBdbyz+RSzmoa+/t2GU/twgpWwiAIlEEkRijOti6NvUd+A4JC9fhTfu7YB/EYPcvwSL
Pvj3ERAS6J2NcE4YgP3pjZj1tmew2p+QxK8xBt4ibeuZmEEgWDCe202D+R1mOYRCeP/UExftLbx8
nw1sBcYAJGm7+71ufwJwQfwUl4Ze4EqTCyEMmOvHP2CDICmcwwbswLzLAAXMcoDV4R7zYBEZYwwb
YyCE2OHGFkJwGIbsOA78wL/X9ys/JiJJLMwWRmWuVMqOFeLhOvfE8LMYMMVyaTmsfYFIEHFVEteu
I+DFnXP+zZLA6jCG2OkxiP19WAvYr52LRy4BgJWYmS/DLPAgj6qg950auKPe/Jmg1wAD1zY+j2Di
qhPL7kab990ohH/Z54T3JUh29tswZEDSzqV3DksLF4T18Iee4qEGFyTC6lipfvwAuUNsLAgXXLxP
6zV6w+qX9qTqx8wUUXTE7Bzfu66LIAgfyOvMN3f05ba2Nq5/FgBhBb7dpdTfd8c7hODXxQFrARGA
v9cNSCCF9VjzHYazKAGyAE8HgJqHeoJeAwy8LdXZRkFku9vCRAW+b8bJe8vQXG8ZbGAFYXzuVAlC
wLbSw0MSIMfUxapH2u9MBmw8kpN9G34VwFfakBJA4TW1OMwMISiplIpHazVCS2hpaeELLrjA35bH
Wlg7RSn1z7GWat26dWbVqlXBDg8GTOXq3OIpCxYOCAb0nnbgYaCAa3Hkf8Qg2AcLAI8DQA9WT6jP
rwUGtnWq88i/I280YzcFklKCUDD3B8HXG4W3/yb2jRhnaidXN5l9gkuVQZhmGcWo7QgNoha/BhiQ
ZTYsCLMvnHb4N9pfKjz7WnPAROGgota6vK3vXHjhhdtePyE25pUqv6KVQcEyQKvg/qoE/9kknAMJ
lLkGR91P4LcKUGccQmxC8N/n4tEHVuJY9zA08nmAcxJWG5qwhf91DMx1qjOPYIK6bCx+5aozoWDu
m/auDzXC+c8B9kMxDrW5xrwuCButX3wG5WSVebcet62LYzNQk8INlTDoAPCFnteQFCYi+L7PbPmk
TDa3tLpWwkaSmaUU0lqzNpfNLAeANWvW0EjHALMIodIq/yIAsUVKW+t5cbcSBD/Nq0X3KqWE1tpu
x8HACrMcjdX9V+HIjwbguwi0F4Hv8SAgQRhAuDqJ+FlVh9cDwQRbvQYYeCypO+Izriq+ll6p6tzK
T7a8pXnQOleFxGyrIQuMMx+UDdvKYygKAI4dxbBb1Oet1GlRYmOJ6DMXthy6fEVv4fHXihSuMnDA
8XjsqEQicVT9QVnzQG/YuHEjgOVjzy1xY1PjZ6SUI2SgtRaNjU148cV/bgBw744soSpTrg6/jiOm
JUCzGZYkBCxsCDCVYC7/Ih5ZWAs1fQszP+OA3k6AZ8E3zcHD9ygooaEnQkuvFgPXb3i7XSdWzSn0
SlRnbf7snNjVQs5BfbugOlsACQh6gofQhzDh1qnOo7WFrdVpkAVbl4RXojAL4BM9u5gfvidISkF+
pfJSxa88BcChKKbLYPZKHoHpH9sxoGlwcOBhAg0wWNZda8IgcED8NwDo6enh7cV/AfBKtB5AkHck
IQ8pwcKHfUKADmfACoj3fxNvWnk+Hnv2GrQdJUE37QVXvIzKPy1ooQJEFpr1BK+9OgzMGJmowSMk
ccS8vCUXml+h6nz/Pu96Z4LkeZs5GDfzArAJEhSwffJplCsJkkdW2FgGxNbSlrd6t+h9ZFUKi4/P
3vv1x616+R9/SQGy8C/O77XWIhaLUblc/mk+m/7SznqhR+RiCv7PfCb9yPauKxS2H+4hgK+Fc00D
5CGDMBULvuSN2OuaJ9CXdSAWSeDNFu6vv4UjTwNouQdBm+AXfdgPXoA167sBOYHWglevGom35bAa
IcVGfr4rGiIQxX1JXiVAwoBpF0oaOQZBSSu+shHBApeILMAjxzjWO43WKJgFQRiI3GtxjZRSYuXK
lS4zU/3PjnJorDEuM1N3d7cc77U1U2Il2t7A4PcaWGvAv/4iHr5qA1bzF/FwRxnhclH1PxxOoPsJ
PMsBUQDuugBr/qowy5kodHgVJfBoxuThP3X2I2/x5BJ2zQvNNa/z1BO/MoncYzbtgvRlsGkgR27i
4A/Hb7zrf1biWHfDlPLjDonDK2yHpfAox9U2wkqQZbZWEp169pRDZt248W+r96AUFkopsW5dn1BK
jZi8np6eMSsTiIG1a9dSNpulUamayGaz2yx8kFLWrqHRTi6lFLTW20pmQ1vE4AZihgPhRplwTyhA
PI+3eSsRhHPwwMUrcRQc4EIC9opB0mYEvz0PjyxeiWPdOXswOWaCRjHw1p7aMaQu8xiqNcZd0QIU
7KN7nXJAiFAV2dgIVGCcqgMhBFsBXAwAc/BAkLNHLBOSrrdga0fVLG8pwtjKDo686cwggrU2B2BW
6x4Lf1Ax8vraMXKnxRh1iEH0/TELICIm3JYDapvP2hGt2fL+6y3YBLCCgPf2AAsK+FNpyzrYHzLE
lxjMBixc0ORvYebUOXhg/WvBFPk/w8AjHVdjOLGYR6VT7lrCaxtSRCjY+93KlZPIa+6vSt9xMjCH
TeQ6/Tb4/nEb7vojIyWBgr2pd+8fPL33hkWOEAeHbC2Dhd06iWNb6rT02RgpxImfnvL6D+qN//jl
7tuAZAEYY4wl8NcyKvdylKTMDEAKQcaY3mIy9iUAA7WUSN/3mcEfymTzBzDg1NuSDNhYLOZUyqXV
ea0u7+vrE6iO1QJsABiyuDaTzfVXTSUa4ceW0pU2tPO1XvToWKEkDVgFiJfwyFP7Yua9LsSJBvao
UzDz1lPAVwLoBehdDKQlKO5AwALwII8HzP98C0e+9zw80jdR3PAqMPDozT22E2u0V7rqxKp6oZ8Z
V4ng/0458bQEOR8d2CXVGeyAxCCbQZKyoxaKymKW1FhdTqNtuQRdbcF2B3HgLT9cr00wrLW5FHDb
bpPChEQ8HpdhGMpYLP4OIcTw0cfMcBwXmzb1QhpzEQCEoeM0NMVlEPiIxWIHOY570OihWGvR0NCA
9ZWyA+DytWvXUuRcSHqxuAxDI+Px+InRObFVhpcXi6Fv48tXAni0p6dnm/awBuzVMF/0Qb9JQOxP
oNMC2NMCMJKQIABFmKEQ9iwCTnUgPxeDe+wAgntXovUDs9HzfPZVKtv8P8nAvD2pW/tsm6pzzYl1
EIBndyrm+4f935Zgn6404Fo97ngdV7aRPNlnK8uPf/HuZ6uHgq7lbZAQ/N0hDi+VoP19WGsBsbUn
emxHXVUKW+MI+Wa710Ef133P/PiVSOFaiIYs9xRLxXuDMAjCMBixk4nBJIVgyy8bKatqsvD7KmW6
xxhjgiAAo8hjYPuaIAhcZr4XANatWxeFieihcqk0NQiDIBjcZk6FdUtlCYj1ANDaOjaiqAZsdY3W
PLYCR74d4PkWOMUCMxhwigjXC+CeAFh6Hh55dCWOvWUIQUjA4Q4oXoE8t4BURwqFCQm8h4j+PPXE
4UodUxcuMnUJG6Zus5stv7eSSJTZPBckNxx+9rPPlrenVdfKBP889UTVItxsn90l6WtjJMi39vkY
0Na24aQioLl2ukcZQ+GCvQ+fFxPuFYMcGgbk6LHXM6/hkZ8ZwAoi8tk+Ti8nj25FT6i3COv/k1Sv
Bq/EsW4M4dQKHBmg1Hs+egZHAzCMPLSroagJVttDDHzf1BO5nmG3YmLeevNH/7aCSFTYPhfugIEZ
SgCa/3efWW9wQH+1xHFTlb6EcXqem8mVA1z5zLHr7/l+7VAYvWG+POmoSTGn8hiIpgWwXC+Fx3q3
rX/HxiEhSzCzf77xuW/PApzVowoJdhUTa2dhdMaBiTXak7xHMLEUINqQotFM2o2UXIMC1xA5FEDZ
aDwTjLvnSWwrdGRH2Ia8VVZWvS1ZVaG3TQX0EFXjs1+LC5k0YKZx4wBUw0b9NvjTrevv+UH3KOat
bZgsZskV/X/dZIivdoioaguPYQfzttVpC1AAZjAWvnfatIaTtvjs8Apwqnb4M5q5dvIa3pVnjRdz
qw4LixQgFCAYoHYUTJ2TinU1u5YnmPdVsoHtGGplPfOOCLWMYR+bHTixaqrVA1NP/HADydMHOTC7
ACIPgChkZgf2Yg3Yaj3qWBtttQFA7PK1pSD8sgC1BGC2YNpWeuXoWDcAEbA1LomDEsabo4Gvvdoh
kSikJNra2ri9vd3synVr1qzhtra2WgzYbq9oIZVKydbWVqpdV0vyyWazZhSzc2RS1OF8grq7u8Wa
NWto9HOj8dtxmiDbvN+23qO7u1vWPW+H81V7356eHt5WNlrtnuMY9/DYauuwM36SbT0/1d0tW9es
oZ62Nm6N1qT6fqlaiTvonqknRCr0CDswUi+3af+OUKEp2Xj42c+u3kqFrsHBPjD9v+NkGh+OkXh9
mQ2PN+5bU5032+CHx22461M7Qnyo/X7u3od1xYVcOMRhyIAz/I6jVGczSoWOPrMCRJb5JWPQ+sv+
5zYNp3+/CqbNrtjcO6osqmFujQE2gO09b3v37e7uljvHMFtjgO3q/Wrx8u296zjn9zUFARStCe/M
WjtjFerXh1XG9ErXZWhZAHKbErharHB/OGv+ZOG+YRP745a+UdiISjYccgQvqhX/Y7tJCAUGQEY4
V5U4/BKBmk1UPzWW6jw6qSMq0BAGbGJC7OvDXgBAzwLk6j3MwLXFU6rzOMcRh/p+sCmXV7ftjEqq
tbZKdc0UAqcZtkcCaCSifgB/gQ1u1Vo/s62NktZd73aEONnY8FBixED0Iglx99Dmvlu11gNjbXKl
lGhvbzcXXXRRYvLkvU+BoBOs5YMsc0IQDQlBT1qIu9Y9//TqVavagx0cMKSUotr9mvfa52QCTmC2
BzNzgpkHpBSPsMFvte74Sz3jMTNlcrkPSOFO5jDsB8xtO2BuVip/ouPI/Xw/eDGfV78bC5DQCvf9
gqgBxlhmsTV2Y3TMWLLsSc/xff+Jzs7Mg9XrO99GJA5mDgMWgkbrbkSWiRwZsHm5S6dvH2tNtO46
GUKcbKw5hJg9FmKdI+Xq/t4Nv9BaDwEgWj31XTxC+vLYUmkMSWxF1Qv9vJtsPmy0BGYoQdD80JST
DyVhHrLg2K44rgAOm4Xr9HHY9ZaXVnfwGLbv2IXoVZX3i3sfsjwm5FeKHIYMckZ7ncfywNuRrVQB
0KZAoPV/Njz70h7GkyYAmDdvWTLZWHqqsbFxehiGKBdL78rn1R9S3d2isLV0IqUUARAQzhVCyPO8
WMz1XBcggjUWQeAjDMNNYRh8k9jobDZrstksaa3tgq6uaXGLG4SQp3meB8d1AQaMCeH7Pizz077v
d3TqzA+VUhQxBlUT14gz2fznpZTzhRCHuq4LKZ0aWyEMQwRBADAeDnyTzec7frYNJh6+Xzab/zxJ
OZ+EONSrux8zIwxC+H7FENEtNqR5Wi94tsqMyrHkPDt58uQZm/r6NgiY/euhdkdJLVZK7cfC/Xsy
kfDK5ZJlExyqtX46ej8GwOedpxqnTHWeb2homBwEIUjQ8M6udyAy10o0G9D78svX5XPqHADIZPM/
nzR58keKQ0MQQmI0mlINXHDzwObH89n0m+ql6sKFnfvFkvIGQfQ+z4vBcar5VnVr8vcgDOfnVcfN
YuwkDh4ztXJUOSGFzABTS1gqT+pGSnbX6fwF9BAA9im8Mk4ysSuOKwasS0IOmOCFyUHicgUlsJMx
xSgJg6z1VpStGQJIWq76Z0dXJI229Ue1UjUu0V7Smq8C4J492MBNKSUBcGNj5WPxRGL64ODgEBFZ
CPoSAEahMNY1pLW2THJZQ0PThQDcUqk0uKm//9b+Tb3XDWzefF+lUkEymZzMoP0jG1EAwIq5K2Ku
b34WiyVOs9ZicGho7aa+vh9u2tT3ncGBwSeYGZ7rHSyIJgHgKOGDUqmUICJOq9y1iWTyOgCHEgkU
i0NDA5v77x/o33zHwOaBB4ulki+EAAMzEw2xmzPZnNZa21RqRCtaSqVSAiCkde66WCJxHYBDpRAY
KhYHNw/0/3nz5v47BwY2P1oulyCEkA0NjR9lCm5XSjUAoHXr1hGB+sulkkFV49he8Q5bOJ+Lx+Pe
4ODQoOfFBZM8J9oSw/s3kWhkMPeVy+Ww4lf8SrkclCvlsFwum7qfoFwph37gl8vlcggSQ3V7d6hc
Khnf9yvRddXro5+KX/HL5XLIjP76w3v58uUJx+VbYp73vuqaDLzQ37/pB/39m747ODj4FAB4rvsG
tmgCwGJEjjBvrwh+JD4WgymANTEhGgJb+WI7CqYdMAoQ3Wj12lEw90098T+ahTxt1x1XzHFIYlDm
sN7bNp+E34ud9W5qwKaQEqv6ep4zxN91SZAFm7EqkuoPpzFIBmxZQMz+6OSDDixEKYZ7iIetUkoY
2PPZWhZCNFQqFZJSfCSfzx9YKBRMfb507cRetCh/MIjOL5WKhhnryGJWPpv+SC6rzsnp9Ntg7ef7
B/qvz2fTZ7e3t5vW1lbSWtsNLf2nxxOJt1cqZWMs/yGsiDfndeZTeZ05ixAeba1ZOji4+dJcNn2N
UkoUCgWTSqVEoVAwmWxONTY2zSkVixUiQhD4Kx2BY3I685acTp+Sz2Xe7Ao+LgyCHwlBKFcqfjLZ
kMlk8+cUCgWT6u6WkV0e3U8vbWxo+ny5XK5EaCRXC5bH5LOZt+azmZMFm2OI8U5r7J8HBwfXMfGX
enp6SgAwffp0ZrAkIgnmbe0z0lqbecuWNYDw+TAIIKVoDAKfAZyplGrWWpva9mpsHCwRzPscwUe6
wj3ahnykYJxKRGE8Hpcg3EssjxAsjnKFd3RQCWeGnlgyvD7ERESSiBwYaq99V7A4Mvo5ig3PlAjP
qB3eWmvbt3noY4lk4thKpWKssasFm2Ny2fSnc9n0mcTh0caY5YODA1/p1B03KaWEM2z/8kjIWB4j
PjraU40tiBaZb0w5arOg4Nvnb+gZBHr8h6a9t8Fa/4qQdi3jCmCTJEcO2OCBN2+Y9t1IJR+XF7g1
soUF0fKKtWcDFKtmdVcxVu2Iw2mbJwMxEHokGirSdAA4Z08U/UfOG5vW+kRPuscHQQBm7raWP9TY
2JAsDg59AUC6HkGjra2NAEB4dIQUwnFdF8ViqdCZy/zvRRddlGhubg601qHWmRsA3LB1a0pxLAAr
pZQ2rCxfsiTzklI3xtvaGoL29vYygEvrD4rob9PRkXuTELJjqDgUep4Xq1T8BXmdvqxepdda20wm
8wiAT2ay+Rc9z/tyqVQyBCxT6vJf6Pb2l2rA90p1HiccccnQ4GDgeV4sqAQXad3x9VH3CwH8Yf78
y96HeGnSUq2fGw+ed8QgYcNQ+aOJROKAYrFYIqJfhGGYSiQS08vlcgrA9UplHa0RRmr+3+rvsbCz
c7MMmYkIAtSv9cInx1rH0Z+5rn08k+l4cudsKDqOma0QUgZheHmn1hvr1qSIqHhnm06skXnQ21at
a2q0iZjBE/JrQ5a/vHzKzH8MsfX/aQcmHSwSrx9iw2IXqo2GMbkEf5VQMNWChfHZnlUpDLlqw9/+
/tm9D/6RR/KsgG1oo/AZj9IqtkPSZ2sF6LMf3fvA5YWXn31id4eVClX1mGFwvpeIcbE4uJ44+SUh
SodbyzNBOFMptbTmvBh53hhLJFHFnaZDAODKK68sRckgzugQR31G9ZZaNGoFcLPWZ5eVUkIp5dRC
OHXXCQBWOJjtep5jwpDLlcrvOnXmMqWUU3tGZEcileqWQAHE4Tzfp5OFEEd4MXdSqVT+TwBXtLS0
SAChhb3AczwIIVy/UvllTme+vnLlSnft2rVm5P1ScunSS/sB9CulBOlx9cuyzEwZlT9PCMFE9Dhx
4ktMpZMA7MPM56ZSqZuy2aypVXjVaTsCgA2Zk3U2rFMXKrJRyI0LY5g5xojE6O+Odj7WY4hX/wao
uib/s701cbguZMQ7rEradlF8kUPrgA5gogNaSGA6xVBiy2IXhC+DTRO5crMNf3b8hrvv7N5Jx9X2
bGFm5/IKzKcZcDjKw65/l51wLhmHhOdzmAVwxu6O+UYS7hAI+UFmSwD9SuuLN6ZV/lZme5QXi70u
qFQ+DuCmmuSK4rWISfmQb0yZAFcIcarKdV4HSd8goodrULT1kqGG3kFcxcTy/SCQUi5Quss6wv5X
Op1+trbR6u1VrbVRSgkGvTsMApZSEgm+isGURXarsE6h0G5mz57taq2DTDa/0nWdbxpjmMDvA3DF
BRdc4IdhmOjfXDoxCAJIKWFA32BmymazvPX9Cib63bAzbRzajQGcE6XjvNUYQwDfovXFGzPZ/G3M
/FnHcY5tbZ05i4h+V/t+XUwXWmu7KJ+3dR4s1llta7+rlXZ2d3ePVZdtMpkOy1UXM28PUYXY3ktE
X/GDwHccqZTuEo6wP0yn08+PtSZOfY3vVuGkMfGxtmbiCJ1SBGBbtiZ4m2guOqBJPljsgseHJYjK
bMtsxQIG6JXgRNak8Pd6n3rsU3sf/FOXxBklNqGtK+TYyWNcBmythPz4B6cccGxh43MP7EYpLACE
THJOPBZLVCoVNowbAZBA+INKBV/1PM+zhHOZ+TsAjNYaNYdQR0fHP9MqrxsaG5eUSkW4jvv5IAg+
m8113S+E+PlQUC60t7c/UwtRtLe3R7a0ua1YKt3S3Nz8kaHBQdd13a4wDL+azXXdw+CbN7284eYV
K1ZsUkqJbDZb23x7MbAfAKr4ftlK8RCBWEGNqR1Nnz69mlhD9s++70NKSSA6SCnlEFG4aFF+X+ny
voADPwg2c0CPEBEzsxmr1nkb2Wc7o93Akp2biCWoXCqVbIgfROO6KQzDzzqOg1IYng/gzrGk6K7k
3tXCoKENf5xR+aLKdop0Nl9LMzXxeNwpl8tLctn0j9esWUPRmtxaKhZ/1djUfOrQ0KDnuu7SMAwX
qFzX3QDfXBzo/+myZcsGaoe+4FGx0LGZGGPiSo1WuytgsT/FKlPJbQ52jXkBsGkUjqiw/dZbXv7d
E0BK7KY2HQRhl/psTYTWwePcCcQASyIpmHcn9A5V7UA1mRmfAQATmjVTW5rvmzt3hdfW1vMUwPcS
ETnSeUsmlzuRiLgmUWuOrbxOXzY0MHgBCfF8FKJwXdd9h5Ty8gY3/te0yi+sxpdVLcOKs9ksDzU3
fHJocHCFEKIfRPA8b5Lruh/0XO/6ln32fWhhWn9Ma20LhYIAgMB1Y8ScABEIGLJFMVi733a9kUSb
bLXDDRhIAIgDAMU4yQyvqjLyUBC4g7tbu4nm6BBB8oNgBpjveuKJR/4xd+4Kb7Cx8U9haJ6wluFI
eWpHLnfoaGfhK6WGhobDmyc1H9M0qfmo5ubmo5ubm49uaGg8dtLkyUcR4cA6DYe11uFQY39qaGjw
aiFogKprMtlz3dM917uxaVLLgx0d6sM1n4QYW9KOZuixY6ejvdceqPJGSnoWJHnXYGbZgZCbTbih
0XUWR0kbrzjmWgBMChA/2PDsQwHzLxwiYQFjxwkIRJEt7JA47cNT9j8xuq/cDaEjWOF+Kh6PT7PW
gklceeGFF1auuurCSnt7wUgSV1hr2XEcwOD8eqmyJYFDiVwufZUNykcYa88IfP/HFb/Sa61FGIbN
TU1NXWmVX1iV2t2ylg995bx5Ja06vmxDMTPw/bnlcvkOv1KpGGMA5gMTifhPlMqfUsuOcoOgAqAM
Bpg5KZI2GTHwds9rCtAstnRaK0c/kMYtgSioMXYyWUzsTgauOfqscGbHYrFEGIaWQJcXCgVz1VUX
Vq6cN68E2G8IQXBdNy6smF1/3W7IzKHBwcGVmzf3Zwc2D+Q292/Wm/s368HBQdXX29cJS78f1Xqb
ll+yfEirjvP9sj0qqARf9iuVOyu+74dhCABv8OKJnyuVP7HOicU7VJ23ZugtOdMA4MPYmWgKkiQb
ffAuBkvZJsmR/TZc3Lr2jpd/h1nOu3czrhIDlwXMH7FgwbuqHIEAK3IATnqlRf/ZbNZUTRl7rrWW
gyAwgnBCJps7PAJzDw1zc+j7oRDCkY7zoUX5/Bu60um/1ydFaJ1lpeBorTcD+DGAHyulZoRGdkgp
zy0WhwyIv6qUulbr9t4tGV9KtLW1UXt7+3MAvgngm6qra6YNzQoQzSKALHEawO3MTO3t7f1vbJv5
TwAtnucl/DBsZebnCoUCbS/2SpKO9jwP1lgm4OnIq0yDgxtfSjZNWs+M/R3pTA59PlQptbG9vUAY
wzyJxiB2hKZZi5G3t7dbpdRkyzgzCAIOjfGZ+GNplf8AEUsLhADPqPi+raLf82eVUl3t7e39O5PS
uL36s1qyB7Fzuc4u+MeOsui2ZIkNr8nTAFYAWJHL5Y4xjKsExNsdR4owxKUA7nK29kBv3/7lMWBa
fTD2gVc8QMTjwa4zr4mTFP0cPuXHKysZEKgWJuweL28Uo9a9z9z3kZYDfu0I8QGfraFxSlCq2sLG
FWLWh6YceKre+OyvdtUW7u7ulkRklMq9L+bFjqhUKsZ1XSeeSJwlhpE0CMyMcrkEPwjCZDweN8XS
OQAurUmJ2kbTGmEqlZKpVAoA0N7evhbAeZls7hjX9d5mrW22Uh4MoDebzQ6He5iZRhVPPKyU+gRJ
9zFm3gug189dsSKWzWaDQqFg0q0zf+c4zkxrDdjaOUT0q5UrVwpmtvUbvk4NZQZ/IQxDuK5LQRj+
FgBWrFjhXXjhhUMZlfuj48iPA0QhBWdpre9VSnmjo3tKKVEFHoWJ8rq3awWtW7dOAgiYnE8m4omp
pVLReJ4XjycS59GwI59grUW5XEIQhmEykZhaLA6dAeDabDa7VT+qXSPTsnLlyufL5bKIx+MjNMrb
99rL1mfX1Q7l+jUBgEwm8+BXlTojLtw1wnITGIekUinpjMWw25bEo6RypDpLoHwkNQqKHEO7qHuw
C6Ky5Y53vPCnUpQyuVtTFmtZVCzsEgv6wCsrEWSQ5VwK+E1hF1Mra15ky5hLRAwiCsLgyXAgHNpq
WxI8AG/yfZ+jkNLi9vb2zbVNPXfu3FhLS4vRWoetra0OAMydOzd21VVXVRi0XggJZt9yBI7X09ND
hULBzJ8/f1KUL81KKXH77X0iSs3sNcYOua7YC4BficeHwzkC4aogqJxvGRzz3I90qNzn5syZc8Oc
OXNQO0DqQh02k80t8LzY2/wgsKZSHgx9+j4A9Pb2mmpKMX2LgZTvV0LX885cpNRPtda31Tyu9feb
O3duc8vUqTN1On3PjuZ35cqV4fTp0x1r+TxjQgYRB36wJvADfyvwa6ABzIeGYchgOre7u/vbqVTK
bA80cBzHfnnOnDnBTvpD7IUXqslEtKm2JjNmzKAoTLfRsi0KIZoA+K2trexY5kEQJU0VeY125NAa
TjeMkDpCWP94agoaSDbtqvRlsN9MrrfZhn88duNdBQYE7YH+ssNSeOMLd32o5XV3ekK+J2QbjqfR
+RYpzNYRdFyp5aCT0fvMb8YrhVOplKyGZbpmksQpxhgAeFKwOWZ0M7Na6IRJ3kFCzIo7zr6VCj4B
4NttbW102WWXTSpXwp8zUWXhws7Pa93xz+jSUOXz7yKmd1trLIjWlRLxf9QSMpRSb3S8xK2ZXOct
ZPZRWs8pbpEEnRc4jtgvOuQeXTVnTlANX6SgdXtPWunLGhubFxWLQxXHcVZlcp2vIxNcrbXeWLPP
F3Z27ucaXCKkvNAPAj+ZSHjFYvHSxYs71kXvH6ZSKdml079PZ/R3G5qaPlsulXxXej9RuvOrxbj7
vWXz5w/U7rdAdc10Ba9whDMrm+tcXi4OXnbZZZf1rlu3jqbNeN2YRSFp3flez3GOiFoj300wp4xV
yXTRRcsTTZOKD1prD3M9b+aaNU+c3N5Ov4lKCl9hpZKdqpSaEu2zkYd9YyNmNDf3zZkzJwTAHbnc
kZ70fpbJdv6YeL+81mcP74VMNneRI51pRCAm+0guq61jiX/cJJzP91o/3FbxwmgnlmGQDwvJVDpe
NAXTKNZYgQ3FLvVGIjFJuF7Rhi9LIc4Zxn+Axp6gSAqTIMy3zPcJIsewDcZfZEGBA0oYMm8G8Jv1
mEXAaoyDgVEoFMBkz3Edr+YC+J7WulyL89aFQSKGy60k0ElhGAaW+RwA325vT9lMNv+jyZP3Omlg
YABOjP83rXO/hqVniPAGtvRhgBpjXgxB4H9t+SWXDCmlxFe+oqZYcu72PG8KV/ji0K7/cFrl7oCg
XmI+lsHvB0DGWAgOl9XG3N7eblOplBSwmcHBgQMaGho/Uy6X4Dlu1rc8J53NPwjmfgKmcMjHup7X
EoYhEvG4VywOXZHLZq6pLxksFAq2qirac4vF4oxkInFKuVz2HMe5OlGqXJxRuYdZiCLY7gc2b3Wk
G5NSIgzDc5PJ5CoAL0+fPp2qtixCgMI6pBKQtecCCAC4YLpe57QdnS1Vne95pbTK3ySlXMzMgQF/
EcBv6pnXcRy2Jqw+h9nsRNefkIGQCT8FnBDVyAzX969yK6FYu3bDBwH8aWFn53QK+R4v7jaDeaEJ
//kfaZW7kwj9AI4jEu8DgDAIGQZfqyZyhHZevxNMbxTOaWZUFdIWmJ1aFZIdbinajBjeKJONe5MH
H4zkrpxLzCixsWVrbiuK8ry3v3jfY1Xpu+caYQ1L4Zef/8vpkw8403Hoao9k8/g1B3ICNmULvgMA
nYTVdvU4QhtRLPb1QsrzY7E4hoYG4Tniv+pqjoeNtFQqZZmZLr7iil8kh0obm5uapxCJ45XKfVxr
+glxfsnAwMAUKZ3jYjE5VUrnM7XjKAxChGGAwYHBr+dymStrdmlTEzYZFhdVyuUl0nX3j8Vihwkp
D6tZrFEl0ZAf+Bd0an1Pbcw1pouqhz6byeYekdK9hIj2SSaT04WU02tb1FgDE4YgIZ4vlUoql83c
mEqlZFTgP7yPI3D6olLqg+WyyBPReULIxmRD48FSiIO5pvEZA2MMwiD8HWzwhYzW/6ibrynJZNIp
Fsv7NE9qlgD8tO56ZzIRP11Kic0DA5vKDd6tUXbaiPmt/V8g/H7gY3FDY6PrOM5/LFSdb9G648+z
Z892AdgwDGXc9RqTySTK5dJe23fHYnIymXSY4TiOjI+V9FlrUrexty8ZefhfZnK+XKn4XY7jTI/F
4m8SUrypfk3CINhcCYIvdebV/Uop4bT3PdAP4IM/mPKWDwYwRzPYMVVgYbKw9QkebGDJZSodLJL2
QErEQmvcl6jCYtfq5kJJ4kUBcf+R6+98OJLI4tXoo1PDPNabnvv++ycdeE/C4Q8w8T7GgsSOJXHk
XUQ5NObXv9z0z79iF0sMjZENjifTxWLRhsa8kNWZZ2t20CiIHCilxHKth5TKnV0uV2YKAVjLQ1V1
MH2XUuqdlvn/GROeak3pcAYnAWwWQjzCxv44l0vfNereFsD3lVK3WXAqCOg9zPZAa60rQBukI+8r
BpXvXab1k2OUAFb106qaumzBgq7vxuP8sZCCkyz4AGs5IYkGiehpEuI3pcH+X1x22WV9NdV9rOSM
6F4+gPlKLVkZwny0UqmcQMB0CxsjEpsF6HGS4laVXvCLUUADloDM0NDg3kKg74knnqhqL2EowtB0
BEHIzNyzbP78gcu/+lXaCmGkatuT1vqFTCb7mUqpfCBJSVIYCQB9fadYYBWcMHzZB+YPDQ06xNU8
aT0qnTOVSllU4VG/PTQ0+IdyuWyrdQ1be7OZmH3fEQ7CJ6OIREBEN85T6pdNiKcCP3iPhTkADAnQ
BinlH/3y4Pe6urq2RCB4D5bHjUeV5n9BV8DdVFW0O+ePduX3YyUdjP5srO+MKusbE0xvRwkNYyXv
j3Xfsb431vtV86fH/357cn5fTRpr7kaPq/47NIxBhJTYB+u3eoHfj3G3k3bjgE/CSZb+hb1jFSB+
DwhgFsZjwwKzcBJW21dS3B85p+pzjcOdgaZpbR3GaaovUKDu7m4RhY9q3WIom83K7eEubflOG9cg
b+oT73cGtmase9TjWo2Bq7VTuF7142ZmqmWDjYWxFXlpR8xjfRhmZ7GyRuFgjX5/qiXe7Oh+Ncyt
ndIIs1mDUYiko9dtvGsyQf/+RPjXaRT0GnmX/+trOUETNEETNEETNEETNEETNEETNEET9P8v45iZ
acJKnqAJ+vchrss2m6AJmqB/Vwm8bNmyhsHBwX1Cx5loRjVBE/RvQk4YUuC6gXPggQeW//jHP77U
3Nw8wcATNEH/JrR582Y66KCDJhI6JmiCJmiCJmiCJmiCJmiC8MpaW0YI8GJHyeZKKTG6cgUjw1Pb
83ETdj7Xk7Zzf9rGs2gnxrCj9x3vmDDOvFXawThpJ95zT45nzHndxnyNZ6y78uxduH91bbexR1/J
/tmZNaBd5J0dvmMdVhaN7oi+wzKu7ZSJ7amAFO2uQ2n04u7M+76aCebV8TC9gjnao0HBbTDta5JG
79NXivG8k6WQ4y4THO99xyoRpfo+O0otmwqnfBhCGtS64yGM3aWdOnJLjo4RNZRK4d8XL+5YN7rL
eT3A2liLrpSK1eE+0UUXXRS/8sorS2M1kT7zzDPj74i9w8xZVQUFU0o5lUqliZqaHM/32fM88n3f
aK37AHBUDtbkeZ4bfd47anxb3nfp0hmiHL4hILOps9qIa/h7c1esiLX09gZjlW6N1em+rqRNRL1p
t0tz586NNUyf3hirVGRjY+PQJZdcMlT/fKWUV6nEmohYeJ5fa31ZjmBjd2o83d3d8u6773auuuqq
yk5sju09b3hMjpM4LgzDYm1/RO882fc8YiHCWLlsfM+Leb5PAAa11sVh2DjUSuaGO82yUioJoNH3
PPZ8v1KJxyVZ63i+zwA2aa3DFStWxIaGhhoHBkaMrai1HtzGgcaXLl68d4OhtjDktVov/Fv94X3J
JZc3Ok4Y9zyffd8jAPA8v6+6X5mUyjb4vpdAI7Bk0aKNtX2jlHIBYKz1jXo5kdbaX7FiRay3bu/U
AyIo1Xk0HKcRoXxS60vW14976dKljZvCMO75Y42rtk6LD/M8Oc33i49orTeJ2sOzua6rhPR/KFh8
gQTSufySu5TKnbYFM0jbdFp/JJdf/FuHzcLABud4cfG9bC5/zUUXLU8A4Ah6BJNb9llGwn2w9v+a
6gWA09n8d5mc39cGtGTJkslNk1uezGRz82sMWFMROjs7px96eOsT/5yx/iwMN0WXZzU0Tb7D9e01
DHlDYOhGC7l47twVHgCsefzxTzte4reVwH6DROx6leu6LZ3W7wTA0SnISn2zMau7rheV8DsMfM6B
zOvOxb9TKn9iDed6Um//twPgLbUFqGkaSnW+re3IY55YsKBrn6gx9bA0Z6IzIZyFo+tTxyqY32vv
qdclrfiRhfxGsRL+KKu7rr/00iV7RYiQYJJLG5vlrW7MXmvZud5xEzcyy7n1alg0r68/Yuabn+zI
5d5Um+faeB557LHTJrfss2x7p32tzadhuayhWdwy/DwvcaOFc36NIbTueifIvdOyPYskZdLZ/C+V
Us0A9rXs3Oj65tteJbyFybnbCey1Xix5E5NzerS/HIA4rXLXpVXuuwBx9TOA2TndiyVvcgJ7LZO8
26uEt7i++TaTcwOA/QFgY2//9aGln7qeWclwrnNiyZsA8dnRkq22xzLZ/Nkx397CwCfIwaq0yn2j
Jr2Uyh4eb6j8yYnZ65icG9yYvT4W5xsAHAIAHbn8IRbiHjdmr3EDe002t/jXaaUvAMCh4+zH5Pxs
7ty5sVECyXth3fqb68drpTymnneUyn9A5xbfAUkdsOEXSPo/Vbrzm9EcQqmuNw2V/D85wZZxuTF7
QwU4qLZ+HZncNST5itCYT1uSv1W53FmOUor+uW79T4QQa63xT49OTGjdeToTDmfmXxFRmNadHyNr
l4LpE1nV8UA0uDiEu7ipufiL2bNnnzp9+nSOIFJiIHrjtGn7nQbglpUrVzpz5swJF+XzByPk9zLh
+ZrUGDAmKRkCwBnLli37ZnsqVZx9++0OgKAS8Kc9l6cD3FiXQ3ZgGAa3E8KFg4ONsQMO8MJ4PG4F
1lahk6x4XcjB3blset5FFy2PNzcPnUpSXL106dITisXi0IoVK2Iv9/b9DxPdSyY4t3aaqlzuU0z0
JiK6K1qbGVKIxFZQR8J+SoL63Jg5C8CyCDsYAGCtmCwE77OTSJwzTBCovM7cq5Tai4Wz1IubxTqr
z42ef0AQVq5+ce0LBWC629bWbHp7S6aGqBgxpGVyP2PZvkyGzgbw1ba2NlGHLd4IEtN2ZjxC0P7G
D656ce3Um4G1I56nlHKMtUsJslNlFt4GAItU/n2iWlz+glIqBcBnOJ8E88cETMqvhB4BfgSVEyq1
5PVMwRvAFFNKHTZnzpynqkwVFvxKeIsAfEtOgYCbCeaHLXu1uL29vUHEJfsGxuqXZky7p+HxQeeg
gxw7/ZCDQ6DaWqYOhdIuyudfB4Mvg53TM+lLn1u+fHli0yC/q+fRRyjq9TmZA94gEaaAg0Rv72Zu
aenltp4eAwAe0BwQ9T+25uFPtLa2SgF+gyXRvUjlH+1Kp+9MZ3ONk6dM+wCAW5VSrtY6AOT7CJzU
VYwuEGhfySJeAxfoULnTGPxNMD6t1aI/RnPazOQuMo6zD4DNgG2xjA2vmzH1448/PuiEoWNbWno5
FgFGPLzm8U8TcNCaRx76ULX1y+JDIOwUh8k5k8GTs5lFH6lDfGClOn4BALlsmpRScWs5TxBnZDIL
HlBKOT09PRypwV9JZ3N3TN//gLN1pmNVbS+HoVklpDgDwC21PStCe7YQ9HPLtM++++7rADBuEDRY
Fn8AUWVwsJQC0U3TlTJKqUYGv6tcKX9HSpEc0RPT2k35qloRjkZkYGYD8AAAvvLKeSUAN2dUfmGx
aKZprf+WyeTOY0ED+UzHAmamnp4e2d3dbYnoByPVHfbrIEBrIHStlulgvzL0YceL37xs2bJvXXLJ
JcUaSoMQCClqE4IdI/r5ILkpwmTqVUpdw3CuqPtGGATBy6tWrQoiVMURamK148CyqRalE4f6zYeS
DfJnSqkZ7e3ta2tzQUIYa3inxkNEgQnNxlVVU2X0NQ4TN5WG5L21diFdRL/BFowxH1rbRRntC6CS
y2mbSqWCQqFgIyjVkCk4lxjXG7BHcM4D8GUANWjZoFAo2HQ2X4a1vs5pq5QKog6FsIAvrb9h1Zwx
xwZguLULOyHtzWTLWl/6HADMmzevBOC3NQkcBAEkOQM6O1INZmYCEYwxTOQMRQeDAfBYWuXudYSY
CeBOZnyPYP6zfl9bok8z8/e2wFGyT9ZyTbt9Yd1LS4jpc0ot/GMd72wGML9+RxBhYM6od6ytpQQd
ZAmP1Q6syCz4mwPGxwTwXzXIDq3bh+FI1q1bJ1etosBafSwJbNK64y81PN+6m1uwvJGtTQFYFa3v
XpL4+8x8llLqbXPmzLnvkksuaSKId9jQpklQV0tLS2SzxSxRmCDQ1ZZ4PoCbtNacyeY/BvBTUspn
mLHf8EaTVHGlO3ORyr+PwPFkMmkrxcpzPT0PraltfBAdltb6rcLKinRoTmjs0+DKM5HS835mvi7q
uCcKhUJIRDUIFN6yoUc4aUR1EzpzCHzP4sWL16Wz+Rf6B0ufBHBdrc9tpB7TTjIMEeGEdC4XE+RO
siZYyswr61ANbTyWOCmd1gDgJRrjPOT7j3Sl08/Onj3bWbVqVWBR/hyYHlm+XK9Pq9yDTPIcAHp4
PNYSkaSddPaxG3Pfk05rp/a8SpEfBkovaK3LmWzu5w1N4W9VrusyZLO3aaBcO+xSbW1UiN6p9v4R
nAxprc3SS5Y2DbF/AhAulYBgcs5XSjVrrQcAUB30jIjugba2Nspms7U+fyRF/NRFKr+/sNZNJOLM
HN6/aNGil7Z0pdARPG3wKOA8n80v/omE+Ho6fek99b2OpJQBDPbrULnTpISQcBBwMJTNZu8GEEop
TWB577Re/HamsBiT3tv8IHg7bLA4grr5iYVz7sLOzum6o2Pdgq6uaQjMIUHZ+2LN3CACoapV4oUX
XpoJCV/r9O/reacGOVQoFCItgkIC75fW+lQAjksuG2MGe3oevjva1t8FyV92Lr5Ml4v+Dzo7M08A
gGDiBmv55Ww2y7UepTU1raYSw5FTmLh/tOs7OkkYsBuBLc2PCRBW8BBJdFvI2QA4mWz6GIOfFMI+
AiDZ09MTPatimTApm+1YDUCk0/rtEQjmGTbEDQB7FPWZidTCgMFvFOAPEfCBIAw/wmyP7O7ujqSl
9V3HbYOlzzPxHxh4fmhg01lRDyIwISZBvVpr3jKGqiq2rQ572Ww2uEipFra2Vav05QDAIWeJ+WMA
aObMmWYcvZCGN6t0xClkxWJYexMzVKfOXDdsOxMbgI4nSR+CoFNDYz8srDgAAE455RSrlPIIfJJo
DnPVE8Zk2eKkFStWxO65556dbgeSGoFjTG8hKYafBwQH1FqY5rIZZZiXSSnPkG78TqVyn9vCNNtt
3MbFZPAJMP1Oa71Ra72eGb+3EJ+KnFhyZ9zslvgEAa6OzdoP+z6mjW6qVuvu91jr4Z8A+HcMe2mu
c8nv01rPquFXMRAS8RQCPmQNTrPg04nppBkzZhAA+D5CADME7OcFi9uMMe/nkD+utX5uzpw5jtZ6
EwH/64T2UwDI8cNPgvHg0qWX9tfs+khNpMg02RsWm6NDbQQa5kh43SBkYB9YOp0NnWbBp1vgpOga
0lo/I9i83xiOxeLeNUp3dnd1dU1zwFjHkt4YNbqSUSdztLe3i9ppwWT/JixNj75D0cnLra2tEoCx
sG8EsL6usZMRlvYCzM+JnIvnzZvXwMCHIJylsMYDbQUEJ4iIM9n8zZDi4ws7O59BaOOdnZnHMtnc
GSPR7EVj4PuFzly2q/4GL774vAsgIJINZb9ya5dWHWmVc4LADC1fvnzogAMOiAGoEKjXwh4G4M5o
/JaZkc1miYhszYk0GvY0rXS74zqvVzr/5bTSCSIEUjhvWag6j3/3u999f73krmkzo/rq1hiYtdZg
Ag8F5QVv7mn7+yOtj98n4T1aD8JGgBeY4rJ8Nv/rsZpVp3XnBxwhjjQDODOt8gkmLjmObN3Q239q
oVD4+egwWl9fn1BK8ejx1HXCdUxQXprT+nZs3V/XKKWEVh0/BfDTjlzuSCLcnE7rx7RWf1y5cqW7
ja4UUTtRni0dejKTzc0nImbQdGaaBeDa6DtiG4ddda4AhJVwUWfncKQAY3W3r0UXutvbLQFXA7ha
dy45naxYuaCr64QlixZtIOYkIB7O6/R5Y4VoiII4sfuYznZ8QancRy3w2c7OzFOR6msjfvgumHIA
loPog4JJA8DatWuH59WKau88Y/hp4dA0pZTT1tbGNd6p9aWqqcTWegmS9q85nTlvO03gXwBwaVVb
nLmsEtibhABdKxhnz5u3rCFy6DARcaFQMAu6uqalUinZmcmsAbAprfQXtdZhNGmstfarjZrFF1H1
GtZm0TJRTGvtM+POxkl7Xc1EpbxaeH8YOpNGN4wipkApJYjDHzP4ECfkKwD6QSTMYUciP1ohxFYL
3tfXZ2s2sCSKdXd3SwGzkAR9ZVE+/4YLL7ywEoHJ3wiiixSUqH/fSJrsWzcqK8SWBoYEcYYQzveI
iUS1K2sIolslcF596MtaLmutrdba11rbbaExEsO4VkxvL7QbAfEtJv+HWmu7bt0MHp5DK1xso58S
WXuOIPFjZiIhYJmZBMn/EkRf2PJtCSKqaK3tnDlzgu2Nh5nZjvG80YzCzKIzk3kEjF8LIY8HgEcf
fVRUpY0Y7rzT29sio8jFCUzsMuj+2oEiiP4ECy+t9Sytte2tqvy1tR1jfGSl3PbYxpDYWxqidSz4
BRO/5ITh60d2BBn7/VzXtbWDUuvMzUxoymTzX4ps9er4rb0XQLlDqc8RqAyEf6gDjAdA1on+3dmZ
eQrAWgtxcXt7u6njHauU8pRSLdW5s7w9iOf6g6pQKJhyeVAD9AZH6/RdmWzuxqZJ/p1ady0sl/Fo
MmnjhuXnbGAOaW1tPVMpJYzkc4QRP8vmljQFFHbbYrHsxONvJJKarS3k1aI7Vq5c6UZGeLJ2qgo4
1zqSng6N+QQAOA1S2sA0jLQvOdnS0uJeeOGFmzqyuR4CvjTQ1PCFmmlGYrivLJhZSkccls7ljhWG
ElZII6w1VdsHRVQbRzuRyvRiWnUuk8beMHv2ylOmT1/LWmf+O63y76Fc7PY0d6mwPPikSCYbHUtf
smwb8tn0F6ttXThpLccAIJPNf5LBJr1ofnZU/NV7U+vMh5XqOkLrRY8SkxSOPEwpdZy1IpZIJOzQ
UOWFrq7082PEuJPRQUZad9yQzuY+05HJXdSZm3NlzZPvOrFjlOp80VoTcxKuDYphWevMg2mtTwDT
/umOS//f6IVOZ3MPptP65Hxe3VHV4Oj1SnUeB7AXi8V4aKiyvqsr/ffR4yGimOM6b1aqc0PtebbC
Q1ovenjevHkNDY2TvieFuDGbzd6d1l3Hwph3G8deBQCVyhE2UhtdUX0vtLT0MjNIZWmRZSzPZhZ9
v36cHSrXL6xYxIy7stne2jiS1toxGJWT5Ii3prWWgOskXAfFsLypM5N5bHSeQkcud7iw4gpH2Eyx
6D7jeWE7GC6xWVPtFmFZEu2fTuu3Oo4rrLUci3k0wMHflixatCFgdkCUaG9vN1Wbd8m5FuHtSi35
VXt7+9OR59nPZPI/SSSbri+Wimd3Vpmrru8RJ61lOSzZ2ZkNhP+dzS122dD3AKdohX+kZZMVkhSA
OyN+2V+pzrcA7ABk68eVyebPFoR3sPUUULRMzhIAq53IvulSua4niHBxLM5eyCQl+BHmcG6tAVWh
UHhKXX75KVQOc9LSTW680YLgM3BtLpv5UbWz3e1VVHrwXcT0dPXkWPBMJpP7MpGpqmaVymYieVvN
HvB9b8CNmV9HneoIAt8lg6eqjZcBAX7IUJ2KJex9UsaPtpXKfBYkCVZYoCSACwAUBfhRhm2oxQjz
uuPqtModMX36ywdrrZ+MVNCvqFznWQ7JDhlvcJghiXC/sGaeUlnSGszAnSB6MfIYz4AUX6slRkQO
OFHVQI74lhV0OIBHWfJDUsgT/VDMJylgwDHp4ocAfphKpUShUDDDPXtAt5NwNtSYiAW+TIxzan2R
mO3vSHinsLBHgoQkAykcehbAg9aIQ6TAikjtcwDY3t5e2dLSErC1XwPJ1wO4gw0/CU/2sQ0vBREb
5phw6DYA34pUOG6t9f6x/Hu44mRme1TteVbYvwG4pLGxsQTguww6Xzrxc4ltzDBf3JlRT1THsNZU
5Q4/C0l3b0l2UM0Melwi/LlSylm3bh0BQNW3Yn5iyDkmm1VNw8kpjLut5WdrmkatTSoBd0oSHzAs
3kMckrHkkqEHAOTqnViRk+lpJtxuyVmWaERgAmZr8PnOzmrSB4XhP8l1/g6B+YYNALYhW8/z7eUA
NgTW9jksbweAqA/xk0rnv84cvhPAP2pMShT+Igj87xtf/jJ6X8PMiFT+O2V1bRH1+X1mQVfXu+MG
eUh7I9i3kuCzxVKdTt8ZOameIzj/4Kojl8DWGmtdx4SXAdhAHPsVRHAcyfA7oBgxm8cG+nq/+v8B
IwYBnkrO/k4AAAAASUVORK5CYII=
"""


def _resolve_logo_path() -> Optional[str]:
    """Decode the embedded Lee logo to a tmp PNG file. Returns the path, or None on decode failure."""
    try:
        bytes_ = base64.b64decode(LEE_LOGO_B64)
    except Exception:
        return None
    tmp = tempfile.NamedTemporaryFile(prefix="lee_logo_", suffix=".png", delete=False)
    tmp.write(bytes_)
    tmp.close()
    return tmp.name

# =====================================================================
# Constants — registries the helpers read from
# =====================================================================

RDU_MSA_CITIES: list[str] = [
    "Raleigh", "Cary", "Morrisville", "Durham", "Apex", "Holly Springs",
    "Wake Forest", "Chapel Hill", "Garner", "Knightdale", "Wendell",
    "Zebulon", "Rolesville", "Fuquay Varina", "Fuquay-Varina", "Clayton",
]

# Named markets the broker might say. All map to the RDU MSA city list for V1.
NAMED_MARKETS: dict[str, list[str]] = {
    "RDU MSA":        RDU_MSA_CITIES,
    "RDU":            RDU_MSA_CITIES,
    "Triangle":       RDU_MSA_CITIES,
    "Raleigh-Durham": RDU_MSA_CITIES,
}

# property_type values exactly as stored in Dealius. Note the misspelling on
# medical office — that is canonical, not a bug.
PROPERTY_TYPE_MAP: dict[str, Optional[list[str]]] = {
    "industrial":     ["Industrial", "Flex Warehouse", "100% Warehouse"],
    "flex":           ["Flex Warehouse"],
    "office":         ["Office"],
    "retail":         ["Retail"],
    "medical_office": ["Medcial Office"],
    "lab":            ["Lab Space"],
    "land":           ["Land"],
}

CANONICAL_COLUMNS_LEASE: list[str] = [
    "comps_id", "comp_name", "street_address", "city", "state", "zip_code",
    "county", "property_type", "space_sf", "square_feet_sold", "building_size",
    "acres",
    "lease_execution", "lease_commencement", "term", "lease_type",
    "free_rent_months", "ti_allowance_per_sf", "asking_rate_per_sf",
    "effective_rate",
    "tenant", "landlord", "lead_broker_s",
    "landlord_rep_agents", "tenant_rep_agents", "link_to_comp_profile",
]

CANONICAL_COLUMNS_SALE: list[str] = [
    "comps_id", "comp_name", "street_address", "city", "state", "zip_code",
    "county", "property_type", "square_feet_sold", "building_size", "acres",
    "property_year_built", "actual_close_date",
    "asking_price", "sale_price", "price_per_sf",
    "asking_cap_rate", "actual_cap_rate", "investment_sale", "off_market_sale",
    "buyer", "buyer_dba", "seller", "seller_dba",
    "buyer_rep_agents", "seller_rep_agents", "link_to_comp_profile",
]

# Display layout for the Comps sheet — (header label, dict key from row).
# Order is frozen and matches Run 1's gold-standard output.
# Lease: space_sf is the primary "Leased SF" column (69% populated). When null,
# format_excel falls back to square_feet_sold (85% populated).
DISPLAY_COLUMNS_LEASE: list[tuple[str, str]] = [
    ("Comp ID",          "comps_id"),
    ("Property Type",    "property_type"),
    ("Property/Comp",    "comp_name"),
    ("Address",          "street_address"),
    ("City",             "city"),
    ("County",           "county"),
    ("Leased SF",        "space_sf"),
    ("Building SF",      "building_size"),
    # "Lease Executed" (lease_execution) removed per broker request (gi-plugins#106,
    # Will Fogleman 2026-06-17). lease_execution is STILL the lease date_col for the
    # WHERE/ORDER BY filter (see CANONICAL_COLUMNS_LEASE + build_query), so it stays
    # in the SELECT — it is just no longer shown in the broker-facing display.
    ("Lease Commence",   "lease_commencement"),
    ("Term",             "term"),
    ("Asking $/SF",      "asking_rate_per_sf"),
    ("Effective $/SF",   "effective_rate"),
    ("Lease Type",       "lease_type"),
    ("Free Rent (mo)",   "free_rent_months"),
    ("TI ($/SF)",        "ti_allowance_per_sf"),
    ("Tenant",           "tenant"),
    ("Landlord",         "landlord"),
    ("Landlord Rep",     "landlord_rep_agents"),
    ("Tenant Rep",       "tenant_rep_agents"),
    ("Comp Profile",     "link_to_comp_profile"),
]

# Sale: square_feet_sold is the primary "Building SF" (63% on sale view); falls
# back to building_size (50%). Format_excel handles the coalesce.
DISPLAY_COLUMNS_SALE: list[tuple[str, str]] = [
    ("Comp ID",          "comps_id"),
    ("Property Type",    "property_type"),
    ("Property/Comp",    "comp_name"),
    ("Address",          "street_address"),
    ("City",             "city"),
    ("County",           "county"),
    ("Building SF",      "square_feet_sold"),
    ("Acres",            "acres"),
    ("Year Built",       "property_year_built"),
    ("Asking Price",     "asking_price"),
    ("Sale Price",       "sale_price"),
    ("$/SF",             "price_per_sf"),
    ("Asking Cap %",     "asking_cap_rate"),
    ("Actual Cap %",     "actual_cap_rate"),
    ("Close Date",       "actual_close_date"),
    ("Investment Sale",  "investment_sale"),
    ("Buyer",            "buyer"),
    ("Buyer DBA",        "buyer_dba"),
    ("Seller",           "seller"),
    ("Seller DBA",       "seller_dba"),
    ("Buyer Rep",        "buyer_rep_agents"),
    ("Seller Rep",       "seller_rep_agents"),
    ("Comp Profile",     "link_to_comp_profile"),
]

# Land sale variant: brokers price raw land by acreage, so swap the $/SF column
# (col 12) for $/Acre (price_per_acre, computed sale_price/acres). Same column
# count and positions as DISPLAY_COLUMNS_SALE so format_excel's positional
# index sets stay valid; the "Acres" column (col 8) is retained.
DISPLAY_COLUMNS_SALE_LAND: list[tuple[str, str]] = [
    ("$/Acre", "price_per_acre") if key == "price_per_sf" else (label, key)
    for label, key in DISPLAY_COLUMNS_SALE
]

ASSET_TITLE_MAP: dict[str, str] = {
    "industrial":     "Industrial",
    "flex":           "Flex",
    "office":         "Office",
    "retail":         "Retail",
    "medical_office": "Medical Office",
    "lab":            "Lab",
    "land":           "Land",
}


# =====================================================================
# Validation
# =====================================================================

def validate_request(parsed: dict) -> dict:
    """
    Apply defaults and surface gaps in a parsed broker request.

    Args:
        parsed: dict from the model's parse of the broker paste. Must contain
            'asset_type' and 'transaction_type'. May contain geography,
            size_range, date_window, target_count, min_price, notes, plus any
            extras (preserved unread).

    Returns:
        {
            'validated': dict,           # parsed + defaults; original keys preserved
            'missing_required': list,    # blocking — caller must clarify before SQL
            'applied_defaults': list,    # human-readable; surfaced in email body
            'warnings': list,            # human-readable; surfaced in email body
        }

    Defaults:
        - geography missing → {"named_market": "RDU MSA"}, applied_defaults entry
        - date_window missing → {"lookback_months": 12}, applied_defaults entry
        - target_count missing → 8, applied_defaults entry
        - min_price missing AND transaction_type=="sale" → 500000, applied_defaults entry
        - size_range missing → no default; warning only

    Validation:
        - asset_type unknown → missing_required entry
        - transaction_type not in {"lease", "sale"} → missing_required entry
        - geography references a sub-region we cannot resolve → warning, fall back
          to broader market
    """
    validated = dict(parsed)
    missing_required: list[str] = []
    applied_defaults: list[str] = []
    warnings: list[str] = []

    asset_type = validated.get("asset_type")
    if asset_type not in PROPERTY_TYPE_MAP:
        missing_required.append(
            f"asset_type unknown: {asset_type!r}. Expected one of {list(PROPERTY_TYPE_MAP)}."
        )

    transaction_type = validated.get("transaction_type")
    if transaction_type not in ("lease", "sale"):
        missing_required.append(
            f"transaction_type must be 'lease' or 'sale'. Got {transaction_type!r}."
        )

    geo = validated.get("geography")
    if not geo:
        validated["geography"] = {"named_market": "RDU MSA"}
        applied_defaults.append("geography: RDU MSA (no geography specified)")
    elif "named_market" in geo:
        nm = geo["named_market"]
        if nm not in NAMED_MARKETS:
            warnings.append(
                f"Geography {nm!r} did not resolve to a registered market — falling back to RDU MSA."
            )
            validated["geography"] = {"named_market": "RDU MSA"}
    elif "anchor" in geo:
        warnings.append(
            "Anchor + radius geography is not yet supported (no lat/long in export). Falling back to RDU MSA city list."
        )
        validated["geography"] = {"named_market": "RDU MSA"}

    if not validated.get("date_window"):
        validated["date_window"] = {"lookback_months": 12}
        applied_defaults.append("date window: trailing 12 months (no date window specified)")

    if not validated.get("target_count"):
        validated["target_count"] = 8
        applied_defaults.append("target count: 8")

    if transaction_type == "sale" and not validated.get("min_price"):
        validated["min_price"] = 500000
        applied_defaults.append("min price: $500K (sale junk filter)")

    if not validated.get("size_range"):
        warnings.append("Size range not specified — query will return all sizes.")

    if validated.get("min_acres") and transaction_type == "lease":
        warnings.append(
            "Acreage filter on lease comps: only ~16% of lease records have acres populated, "
            "so this filter will drop a lot of comps that may actually qualify but lack the data."
        )

    return {
        "validated": validated,
        "missing_required": missing_required,
        "applied_defaults": applied_defaults,
        "warnings": warnings,
    }


# =====================================================================
# SQL composition
# =====================================================================

def _sql_quote(s) -> str:
    return "'" + str(s).replace("'", "''") + "'"


def _date_cutoff(date_window: dict) -> Optional[str]:
    """Return YYYYMMDD string for the lower bound, or None if no window applies."""
    if not date_window:
        return None
    if "lookback_months" in date_window:
        from datetime import timedelta
        cutoff = date.today() - timedelta(days=int(date_window["lookback_months"] * 30))
        return cutoff.strftime("%Y%m%d")
    if "from" in date_window:
        return date_window["from"].replace("-", "")
    return None


def _resolve_cities(geography: dict) -> list[str]:
    """Geography dict → concrete city list. Falls back to RDU MSA on unknowns."""
    if not geography:
        return list(RDU_MSA_CITIES)
    if "cities" in geography:
        return list(geography["cities"])
    if "named_market" in geography:
        nm = geography["named_market"]
        if nm in NAMED_MARKETS:
            return list(NAMED_MARKETS[nm])
        return list(RDU_MSA_CITIES)
    return list(RDU_MSA_CITIES)


def _date_as_yyyymmdd(col: str) -> str:
    """Return a SQL fragment that converts an MM/DD/YYYY text column to YYYYMMDD."""
    return f"(substr({col},7,4) || substr({col},1,2) || substr({col},4,2))"


def build_sql(validated: dict) -> dict:
    """
    Build a parameterized SQL string against lease_comps_safe / sale_comps_safe.

    Helper does NOT execute — Cowork sandbox has no MCP access. Caller passes
    the returned 'sql' to the MCP read_query tool.

    Args:
        validated: output of validate_request()['validated'].

    Returns:
        {'sql': str}    # ready to pass to read_query

    SELECTs only CANONICAL_COLUMNS. Filters by property_type taxonomy, city list,
    date window (MM/DD/YYYY → YYYYMMDD conversion in SQL), size range, and
    min_price (sale only).

    No auto-expansion. If the result count is below target_count, draft_email
    asks the broker which dimension to widen — the broker drives, not the model.
    """
    asset_type = validated.get("asset_type")
    transaction_type = validated.get("transaction_type")

    if transaction_type not in ("lease", "sale"):
        raise ValueError(f"transaction_type must be 'lease' or 'sale'. Got {transaction_type!r}.")

    types = PROPERTY_TYPE_MAP.get(asset_type)
    if not types:
        raise ValueError(f"asset_type {asset_type!r} has no taxonomy mapping.")

    is_sale = transaction_type == "sale"
    view = "sale_comps_safe" if is_sale else "lease_comps_safe"
    date_col = "actual_close_date" if is_sale else "lease_execution"
    cols = CANONICAL_COLUMNS_SALE if is_sale else CANONICAL_COLUMNS_LEASE
    # Sale: building SF is in square_feet_sold (denser), falls back to building_size.
    # Lease: leased SF is in space_sf (denser for leases), falls back to square_feet_sold.
    size_expr = (
        "CAST(COALESCE(square_feet_sold, building_size) AS INTEGER)"
        if is_sale
        else "CAST(COALESCE(space_sf, square_feet_sold) AS INTEGER)"
    )

    size = validated.get("size_range")
    date_window = validated.get("date_window") or {"lookback_months": 12}
    cutoff = _date_cutoff(date_window)
    cities = _resolve_cities(validated.get("geography", {}))

    where: list[str] = []
    where.append(f"property_type IN ({', '.join(_sql_quote(t) for t in types)})")
    if cities:
        where.append(f"city IN ({', '.join(_sql_quote(c) for c in cities)})")
    if size:
        where.append(f"{size_expr} BETWEEN {int(size['min_sf'])} AND {int(size['max_sf'])}")
    if cutoff:
        where.append(f"{_date_as_yyyymmdd(date_col)} >= '{cutoff}'")
    if is_sale and validated.get("min_price"):
        where.append(f"CAST(sale_price AS INTEGER) >= {int(validated['min_price'])}")
    if validated.get("min_acres"):
        where.append(f"CAST(acres AS REAL) >= {float(validated['min_acres'])}")

    sql = (
        f"SELECT {', '.join(cols)}\n"
        f"FROM {view}\n"
        f"WHERE {' AND '.join(where)}\n"
        f"ORDER BY {_date_as_yyyymmdd(date_col)} DESC;"
    )

    return {"sql": sql}


# =====================================================================
# Excel
# =====================================================================

def _asset_title(asset_type: str) -> str:
    return ASSET_TITLE_MAP.get(asset_type, asset_type.replace("_", " ").title())


def _geography_label(geography: dict) -> str:
    if not geography:
        return ""
    if "named_market" in geography:
        return geography["named_market"]
    if "anchor" in geography:
        return f"{geography['anchor']} + {geography.get('radius_mi', '?')}mi"
    if "cities" in geography:
        cities = geography["cities"]
        if len(cities) <= 3:
            return ", ".join(cities)
        return f"{cities[0]} +{len(cities) - 1} more"
    return ""


def _to_number(x):
    if x is None or x == "":
        return None
    try:
        return float(x)
    except (TypeError, ValueError):
        return x


def _to_int(x):
    if x is None or x == "":
        return None
    try:
        return int(float(x))
    except (TypeError, ValueError):
        return x


def _describe_size_range(size_range: Optional[dict]) -> str:
    if not size_range:
        return "Not specified — all sizes returned"
    return f"{size_range.get('min_sf', 0):,} – {size_range.get('max_sf', 0):,} sq ft (leased SF; falls back to building SF if leased SF null)"


def _describe_date_window(date_window: Optional[dict]) -> str:
    if not date_window:
        return "Not specified"
    if "lookback_months" in date_window:
        return f"Trailing {date_window['lookback_months']} months from pull date"
    if "from" in date_window or "to" in date_window:
        return f"{date_window.get('from','?')} to {date_window.get('to','?')}"
    return "Not specified"


def _describe_property_types(asset_type: str) -> str:
    types = PROPERTY_TYPE_MAP.get(asset_type)
    if not types:
        return f"{asset_type} (no taxonomy mapping)"
    return ", ".join(types)


def _compute_stats(rows: list[dict], is_sale: bool = False) -> dict:
    """Compute summary stats from row dicts. Tolerates string-typed numerics."""
    if not rows:
        return {"count": 0}

    def _nums(key):
        out = []
        for r in rows:
            v = _to_number(r.get(key))
            if isinstance(v, (int, float)):
                out.append(v)
        return out

    def _median(xs):
        if not xs:
            return None
        s = sorted(xs)
        n = len(s)
        return s[n // 2] if n % 2 else (s[n // 2 - 1] + s[n // 2]) / 2

    if is_sale:
        sale = _nums("sale_price")
        ppsf = _nums("price_per_sf")
        # Building SF: coalesce square_feet_sold with building_size to match display.
        bsf = []
        for r in rows:
            v = _to_number(r.get("square_feet_sold"))
            if not isinstance(v, (int, float)):
                v = _to_number(r.get("building_size"))
            if isinstance(v, (int, float)):
                bsf.append(v)
        return {
            "count":                len(rows),
            "avg_sale_price":       sum(sale) / len(sale) if sale else None,
            "median_sale_price":    _median(sale),
            "min_sale_price":       min(sale) if sale else None,
            "max_sale_price":       max(sale) if sale else None,
            "avg_price_per_sf":     sum(ppsf) / len(ppsf) if ppsf else None,
            "median_price_per_sf":  _median(ppsf),
            "avg_building_sf":      sum(bsf) / len(bsf) if bsf else None,
            "total_sale_volume":    sum(sale) if sale else None,
        }

    eff = _nums("effective_rate")
    ask = _nums("asking_rate_per_sf")
    # Leased SF: coalesce space_sf with square_feet_sold to match display.
    sf = []
    for r in rows:
        v = _to_number(r.get("space_sf"))
        if not isinstance(v, (int, float)):
            v = _to_number(r.get("square_feet_sold"))
        if isinstance(v, (int, float)):
            sf.append(v)
    return {
        "count": len(rows),
        "avg_effective_rate":    sum(eff) / len(eff) if eff else None,
        "median_effective_rate": _median(eff),
        "min_effective_rate":    min(eff) if eff else None,
        "max_effective_rate":    max(eff) if eff else None,
        "avg_asking_rate":       sum(ask) / len(ask) if ask else None,
        "avg_leased_sf":         sum(sf) / len(sf) if sf else None,
        "median_leased_sf":      _median(sf),
    }


def format_excel(
    rows: list[dict],
    validated: dict,
    output_path: str,
    applied_defaults: list,
    warnings: list,
    last_sync: Optional[str] = None,
) -> dict:
    """
    Write the canonical 3-sheet workbook. Layout is frozen (see SKILL.md "Output").

    Args:
        rows: result rows from MCP read_query. Each row is a dict keyed by
            CANONICAL_COLUMNS (subset is fine — missing keys render blank).
        validated: validate_request output.
        output_path: absolute path to write the .xlsx.
        applied_defaults, warnings: surfaced on Methodology.
        last_sync: ISO timestamp of last mirror refresh; on Methodology.

    Returns:
        {'path', 'summary_stats', 'sheet_name', 'row_count'}.

    Sheet naming: "{Asset Title} {Geography} Comps".
    Empty result still produces a workbook with headers and an explanatory
    Methodology sheet.
    """
    asset_title = _asset_title(validated.get("asset_type", ""))
    geo_label = _geography_label(validated.get("geography", {}))
    is_sale = validated.get("transaction_type") == "sale"
    # LAND sale comps price on $/Acre, not $/SF (broker request, gi-plugins#28).
    is_land_sale = is_sale and validated.get("asset_type") == "land"
    txn_token = " Sale" if is_sale else ""
    sheet_name = f"{asset_title}{txn_token} {geo_label} Comps".strip().replace("  ", " ")
    if is_land_sale:
        display_columns = DISPLAY_COLUMNS_SALE_LAND
    elif is_sale:
        display_columns = DISPLAY_COLUMNS_SALE
    else:
        display_columns = DISPLAY_COLUMNS_LEASE

    wb = Workbook()
    default_font = Font(name="Calibri", size=11)
    header_fill = PatternFill("solid", start_color=LEE_BRAND_MAROON)
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Decodes the embedded LEE_LOGO_B64 to a tmp PNG; cleaned up in the finally below.
    logo_path = _resolve_logo_path()
    logo_available = logo_path is not None and os.path.exists(logo_path)
    if not logo_available:
        # Non-fatal: the workbook still generates, but a silent unbranded
        # workbook must never ship unnoticed again (gi-plugins#90).
        warnings.append("Lee logo asset unavailable — workbook generated without branding.")

    try:
        # ---------- Sheet 1: Comps ----------
        ws = wb.active
        ws.title = sheet_name[:31]  # Excel sheet name limit

        # Logo + title band at top, then header row, then data.
        header_row_idx = 4 if logo_available else 1

        if logo_available:
            ws.row_dimensions[1].height = 56
            img = XLImage(logo_path)
            ws.add_image(img, "A1")
            title_kind = "Sale Comps" if is_sale else "Comps"
            view_label = "sale_comps_safe" if is_sale else "lease_comps_safe"
            ws.cell(row=2, column=2, value=f"{asset_title} {title_kind} — {geo_label}").font = Font(
                name="Calibri", size=14, bold=True, color=LEE_BRAND_MAROON
            )
            ws.cell(row=3, column=2, value=f"Pulled {date.today().isoformat()} · Internal {view_label}").font = Font(
                name="Calibri", size=10, italic=True, color="555555"
            )

        headers = [label for label, _ in display_columns]
        keys = [key for _, key in display_columns]

        for col_idx, h in enumerate(headers, start=1):
            c = ws.cell(row=header_row_idx, column=col_idx, value=h)
            c.fill = header_fill
            c.font = header_font
            c.alignment = center
            c.border = border

        # Column formatting is derived from each display column's KEY name, not a
        # hardcoded position. Removing or reordering a display column (e.g. dropping
        # "Lease Executed", gi-plugins#106) can no longer silently mis-format the
        # columns after it — the index sets follow the keys. To re-type a column,
        # move its key between the sets below; do not count positions.
        _INT_KEYS = {                                  # #,##0
            "space_sf", "building_size", "square_feet_sold",
            "property_year_built", "free_rent_months",
        }
        _ACRES_KEYS = {"acres"}                        # 0.00
        _LARGE_MONEY_KEYS = {"asking_price", "sale_price"}            # $#,##0
        _MONEY_PER_SF_KEYS = {                                        # $#,##0.00
            "asking_rate_per_sf", "effective_rate",
            "ti_allowance_per_sf", "price_per_sf",
        }
        _MONEY_PER_ACRE_KEYS = {"price_per_acre"}                     # $#,##0
        _PCT_KEYS = {"asking_cap_rate", "actual_cap_rate"}           # 0.0"%"
        # Left-aligned: long free-text / name / url columns (everything that is
        # neither a numeric column above nor one of the short centered id/code cols).
        _LEFT_ALIGN_KEYS = {
            "comp_name", "street_address", "city", "county",
            "investment_sale", "off_market_sale",
            "buyer", "buyer_dba", "seller", "seller_dba",
            "buyer_rep_agents", "seller_rep_agents",
            "tenant", "landlord", "landlord_rep_agents", "tenant_rep_agents",
            "link_to_comp_profile",
        }

        def _cols_for(keyset):
            return {i for i, k in enumerate(keys, start=1) if k in keyset}

        int_cols = _cols_for(_INT_KEYS)
        acres_cols = _cols_for(_ACRES_KEYS)
        large_money_cols = _cols_for(_LARGE_MONEY_KEYS)
        money_per_sf_cols = _cols_for(_MONEY_PER_SF_KEYS)
        money_per_acre_cols = _cols_for(_MONEY_PER_ACRE_KEYS)
        pct_cols = _cols_for(_PCT_KEYS)
        left_align_cols = _cols_for(_LEFT_ALIGN_KEYS)

        # Color scale targets the headline rate column: $/Acre on land sales,
        # else $/SF on sales, else Effective $/SF on leases. Blank if absent.
        if is_sale:
            _scale_key = "price_per_acre" if is_land_sale else "price_per_sf"
        else:
            _scale_key = "effective_rate"
        _scale_cols = _cols_for({_scale_key})
        color_scale_col = get_column_letter(next(iter(_scale_cols))) if _scale_cols else None

        data_start = header_row_idx + 1
        for row_offset, row in enumerate(rows):
            excel_row = data_start + row_offset
            for col_idx, key in enumerate(keys, start=1):
                raw = row.get(key)
                # Coalesce primary→fallback for the leased-SF / building-SF column.
                if key == "space_sf" and (raw is None or raw == ""):
                    raw = row.get("square_feet_sold")
                if is_sale and key == "square_feet_sold" and (raw is None or raw == ""):
                    raw = row.get("building_size")
                # $/Acre is computed, not stored: sale_price / acres. Blank (not
                # an error) when acres is null or zero.
                if key == "price_per_acre":
                    sp = _to_number(row.get("sale_price"))
                    ac = _to_number(row.get("acres"))
                    raw = (sp / ac) if (
                        isinstance(sp, (int, float)) and isinstance(ac, (int, float)) and ac
                    ) else None
                if col_idx in int_cols:
                    val = _to_int(raw)
                elif col_idx in (money_per_sf_cols | money_per_acre_cols | large_money_cols | pct_cols | acres_cols):
                    val = _to_number(raw)
                else:
                    val = raw
                c = ws.cell(row=excel_row, column=col_idx, value=val)
                c.font = default_font
                c.border = border
                c.alignment = left if col_idx in left_align_cols else center

        last_row = header_row_idx + len(rows)

        if rows:
            for r in range(data_start, last_row + 1):
                for col_idx in int_cols:
                    ws.cell(row=r, column=col_idx).number_format = "#,##0"
                for col_idx in acres_cols:
                    ws.cell(row=r, column=col_idx).number_format = "0.00"
                for col_idx in large_money_cols:
                    ws.cell(row=r, column=col_idx).number_format = "$#,##0"
                for col_idx in money_per_sf_cols:
                    ws.cell(row=r, column=col_idx).number_format = "$#,##0.00"
                for col_idx in money_per_acre_cols:
                    ws.cell(row=r, column=col_idx).number_format = "$#,##0"
                for col_idx in pct_cols:
                    ws.cell(row=r, column=col_idx).number_format = '0.0"%"'

        # Column widths are keyed by display KEY (not position) for the same reason
        # the format sets are: a layout change (gi-plugins#106) must not shift widths
        # onto the wrong columns. Unlisted keys fall back to a sensible default width.
        _WIDTH_BY_KEY = {
            "comps_id": 9, "property_type": 16, "comp_name": 30, "street_address": 30,
            "city": 14, "county": 18,
            # sale numeric block
            "square_feet_sold": 12, "acres": 9, "property_year_built": 11,
            "asking_price": 14, "sale_price": 14, "price_per_sf": 11, "price_per_acre": 12,
            "asking_cap_rate": 12, "actual_cap_rate": 12, "actual_close_date": 12,
            "investment_sale": 13,
            # lease numeric / detail block
            "space_sf": 11, "building_size": 12, "lease_commencement": 13, "term": 12,
            "asking_rate_per_sf": 12, "effective_rate": 12, "lease_type": 13,
            "free_rent_months": 13, "ti_allowance_per_sf": 11,
            "tenant": 32, "landlord": 32, "landlord_rep_agents": 28, "tenant_rep_agents": 28,
            # long name / url tails
            "buyer": 32, "buyer_dba": 32, "seller": 32, "seller_dba": 32,
            "buyer_rep_agents": 28, "seller_rep_agents": 28,
            "link_to_comp_profile": 42,
        }
        for col_idx, key in enumerate(keys, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = _WIDTH_BY_KEY.get(key, 16)

        ws.row_dimensions[header_row_idx].height = 32
        ws.freeze_panes = ws.cell(row=data_start, column=1).coordinate
        ws.auto_filter.ref = f"A{header_row_idx}:{get_column_letter(len(headers))}{max(last_row, header_row_idx)}"

        if rows and color_scale_col:
            ws.conditional_formatting.add(
                f"{color_scale_col}{data_start}:{color_scale_col}{last_row}",
                ColorScaleRule(
                    start_type="min", start_color="F8696B",
                    mid_type="percentile", mid_value=50, mid_color="FFEB84",
                    end_type="max", end_color="63BE7B",
                ),
            )

        # ---------- Sheet 2: Summary ----------
        ws2 = wb.create_sheet("Summary")
        ws2.column_dimensions["A"].width = 28
        ws2.column_dimensions["B"].width = 16

        ws2.cell(row=1, column=1, value="Summary").font = Font(
            name="Calibri", size=14, bold=True, color=LEE_BRAND_MAROON
        )

        summary_stats = _compute_stats(rows, is_sale=is_sale)
        if is_sale:
            stats_rows = [
                ("Comp count",         summary_stats.get("count", 0),               "#,##0"),
                ("Avg Sale Price",     summary_stats.get("avg_sale_price"),         "$#,##0"),
                ("Median Sale Price",  summary_stats.get("median_sale_price"),      "$#,##0"),
                ("Min Sale Price",     summary_stats.get("min_sale_price"),         "$#,##0"),
                ("Max Sale Price",     summary_stats.get("max_sale_price"),         "$#,##0"),
                ("Avg $/SF",           summary_stats.get("avg_price_per_sf"),       "$#,##0.00"),
                ("Median $/SF",        summary_stats.get("median_price_per_sf"),    "$#,##0.00"),
                ("Avg Building SF",    summary_stats.get("avg_building_sf"),        "#,##0"),
                ("Total Sale Volume",  summary_stats.get("total_sale_volume"),      "$#,##0"),
            ]
        else:
            stats_rows = [
                ("Comp count",            summary_stats.get("count", 0),                "#,##0"),
                ("Avg Effective $/SF",    summary_stats.get("avg_effective_rate"),      "$#,##0.00"),
                ("Median Effective $/SF", summary_stats.get("median_effective_rate"),   "$#,##0.00"),
                ("Min Effective $/SF",    summary_stats.get("min_effective_rate"),      "$#,##0.00"),
                ("Max Effective $/SF",    summary_stats.get("max_effective_rate"),      "$#,##0.00"),
                ("Avg Asking $/SF",       summary_stats.get("avg_asking_rate"),         "$#,##0.00"),
                ("Avg Leased SF",         summary_stats.get("avg_leased_sf"),           "#,##0"),
                ("Median Leased SF",      summary_stats.get("median_leased_sf"),        "#,##0"),
            ]
        for i, (label, value, fmt) in enumerate(stats_rows, start=3):
            a = ws2.cell(row=i, column=1, value=label)
            a.font = Font(bold=True)
            b = ws2.cell(row=i, column=2, value=value)
            b.number_format = fmt

        # ---------- Sheet 3: Methodology ----------
        ws3 = wb.create_sheet("Methodology")
        ws3.column_dimensions["A"].width = 26
        ws3.column_dimensions["B"].width = 90

        ws3.cell(row=1, column=1, value="Methodology").font = Font(
            name="Calibri", size=14, bold=True, color=LEE_BRAND_MAROON
        )
        method_start = 3

        pull_date = date.today().isoformat()
        if is_sale:
            source_text = "Internal sale_comps_safe (Dealius mirror, Lee & Associates Raleigh). Confidential and NDA rows filtered server-side."
            rate_convention = "Sale Price is the closed transaction amount. $/SF is sale price divided by building square footage. Asking Cap and Actual Cap are reported by the listing where disclosed; cap rates are sparse in the source."
            caveat = "Cap rates are populated for ~10% of sale comps; absence does not imply 0%. Buyer DBA and seller DBA may be blank for owner-occupied or single-purpose entities. Confirm against the comp profile link if material."
        else:
            source_text = "Internal lease_comps_safe (Dealius mirror, Lee & Associates Raleigh). Confidential and NDA rows filtered server-side."
            rate_convention = "Effective $/SF includes effects of free rent, TI, and escalations as recorded in the source. Asking $/SF is initial quoted rate. Both annualized."
            caveat = "TI and free-rent values that appear as 0 may be unpopulated in the source rather than truly zero. Confirm against the comp profile link if material."

        methodology_rows = [
            ("Pull date", pull_date),
            ("Source", source_text),
            ("Asset type", _asset_title(validated.get("asset_type", ""))),
            ("Property types in scope", _describe_property_types(validated.get("asset_type", ""))),
            ("Transaction type", validated.get("transaction_type", "")),
            ("Geography", geo_label or "Not specified"),
            ("Size range", _describe_size_range(validated.get("size_range"))),
            ("Min acres", f"{float(validated['min_acres']):g} acres" if validated.get("min_acres") else "Not specified"),
            ("Date window", _describe_date_window(validated.get("date_window"))),
            ("Target count", str(validated.get("target_count", ""))),
            ("Rate convention", rate_convention),
            ("Applied defaults", "; ".join(applied_defaults) if applied_defaults else "None"),
            ("Warnings", "; ".join(warnings) if warnings else "None"),
            ("Mirror last sync", last_sync or "Not provided"),
            ("Caveat", caveat),
        ]
        if validated.get("notes"):
            methodology_rows.insert(-1, ("Broker notes", validated["notes"]))
        if not rows:
            methodology_rows.insert(2, ("Result", "0 comps matched the criteria above. Reply if you want me to widen size, date, or geography."))

        for offset, (k, v) in enumerate(methodology_rows):
            r = method_start + offset
            a = ws3.cell(row=r, column=1, value=k)
            a.font = Font(bold=True)
            a.alignment = Alignment(vertical="top")
            b = ws3.cell(row=r, column=2, value=v)
            b.alignment = Alignment(wrap_text=True, vertical="top")
            ws3.row_dimensions[r].height = 30

        # Windows 218-char path guard (shared helper). Flatten to a CWD basename
        # and cap the filename so a deep/long path can't survive (gi-plugins#7).
        output_path = safe_xlsx_name(output_path)
        wb.save(output_path)

        return {
            "path": output_path,
            "summary_stats": summary_stats,
            "sheet_name": sheet_name,
            "row_count": len(rows),
        }
    finally:
        if logo_path and logo_path.startswith(tempfile.gettempdir()):
            try:
                os.unlink(logo_path)
            except OSError:
                pass  # best-effort; file may have been removed externally


# =====================================================================
# Email draft
# =====================================================================

def draft_email(
    rows: list[dict],
    validated: dict,
    xlsx_path: Optional[str],
    applied_defaults: list,
    warnings: list,
    confidential_reference_unfound: bool = False,
) -> dict:
    """
    Compose the broker reply. Helper does not send — caller uses the connected
    email MCP tool with subject + body + recipient.

    Args:
        rows: query results (or [] for ios/empty paths).
        validated: validate_request output.
        xlsx_path: path to the Excel deliverable, or None for the IOS path
            where no workbook was generated.
        applied_defaults, warnings: surfaced in body.
        confidential_reference_unfound: set True if the broker referenced a
            specific comp by ID/address that didn't appear in results — the
            body inserts the canonical confidentiality response verbatim.

    Returns:
        {'subject': str, 'body': str}

    Special branches:
        - len(rows) == 0: body says no comps matched and asks which dimension
          to widen (size, date, geography). Excel is still attached.
        - 0 < len(rows) < target_count: body surfaces the shortfall and asks
          which dimension to widen.

    The body always reads validated['notes'] back if present, so any broker
    preferences the model parsed but didn't slot get acknowledged.
    """
    asset_title = _asset_title(validated.get("asset_type", ""))
    geo_label = _geography_label(validated.get("geography", {}))
    transaction_type = validated.get("transaction_type", "lease")
    is_sale = transaction_type == "sale"
    count = len(rows)

    stats = _compute_stats(rows, is_sale=is_sale)

    target = validated.get("target_count", 8)

    parts: list[str] = []
    parts.append(f"Hey,\n\nAttached: {count} internal {asset_title.lower()} {transaction_type} comp{'s' if count != 1 else ''}"
                 f" for {geo_label}." if count else
                 f"Hey,\n\nNo internal comps matched the criteria — Excel attached with what was queried. Want me to widen size, date, or geography?")

    if count:
        rate_bits = []
        if is_sale:
            if stats.get("avg_price_per_sf") is not None:
                rate_bits.append(f"avg ${stats['avg_price_per_sf']:.2f}/SF")
            if stats.get("median_price_per_sf") is not None:
                rate_bits.append(f"median ${stats['median_price_per_sf']:.2f}/SF")
            if stats.get("avg_sale_price") is not None:
                rate_bits.append(f"avg sale ${stats['avg_sale_price']:,.0f}")
            if stats.get("total_sale_volume") is not None:
                rate_bits.append(f"total volume ${stats['total_sale_volume']:,.0f}")
        else:
            if stats.get("avg_effective_rate") is not None:
                rate_bits.append(f"avg ${stats['avg_effective_rate']:.2f}/SF effective")
            if stats.get("median_effective_rate") is not None:
                rate_bits.append(f"median ${stats['median_effective_rate']:.2f}")
            if stats.get("min_effective_rate") is not None and stats.get("max_effective_rate") is not None:
                rate_bits.append(f"range ${stats['min_effective_rate']:.2f}–${stats['max_effective_rate']:.2f}")
        if rate_bits:
            parts.append("Quick stats: " + "; ".join(rate_bits) + ".")

    if applied_defaults:
        parts.append("Defaults I applied (push back if any are off):\n  - " + "\n  - ".join(applied_defaults))

    if warnings:
        parts.append("Heads up:\n  - " + "\n  - ".join(warnings))

    if count and count < target:
        parts.append(
            f"Got {count} of the {target} you asked for. Want me to widen size, date, or geography? "
            "Reply with what you want me to expand and I'll rerun."
        )

    if confidential_reference_unfound:
        parts.append("On the specific comp you referenced: that comp is confidential and not retrievable through this channel.")

    if validated.get("notes"):
        parts.append(f"Noted: {validated['notes']}")

    parts.append("Let me know if you want me to widen, narrow, or rerun differently.\n\n— Will (via the internal-comps tool)")

    subject_count = f"{count} result{'s' if count != 1 else ''}" if count else "no internal matches"
    subject_kind = "sale comps" if is_sale else "comps"
    subject = f"Internal {asset_title} {subject_kind} — {geo_label}, {subject_count}".replace("  ", " ").strip(" —")

    return {"subject": subject, "body": "\n\n".join(parts)}


# =====================================================================
# Feedback
# =====================================================================

def format_feedback(
    rating: int,
    what_worked: str,
    what_didnt: str,
    query_text: str,
    xlsx_path: Optional[str] = None,
) -> dict:
    """
    Format broker feedback as a structured payload. Helper does not send —
    caller routes via connected email MCP or writes the fallback file.

    Args:
        rating: 1-5.
        what_worked, what_didnt: short free-text answers.
        query_text: the broker's original paste.
        xlsx_path: path to the deliverable, if produced. Used for context in
            the email body and to locate the fallback directory.

    Returns:
        {
            'recipient': 'david@groundedintelligence.io',
            'subject': str,
            'body': str,                  # plaintext for email send
            'fallback_filename': str,     # 'feedback-{YYYY-MM-DD}.md'
            'fallback_content': str,      # markdown for file write
            'fallback_dir': str,          # dirname(xlsx_path) or CWD
        }

    Caller logic:
        1. Try Gmail MCP send (subject, body, recipient).
        2. Else try Outlook MCP send.
        3. Else write fallback_content to {fallback_dir}/{fallback_filename}.
    """
    today = date.today().isoformat()
    subject = f"internal-comps skill feedback — {today}"

    body_lines = [
        f"Skill: internal-comps",
        f"Date: {today}",
        f"Rating: {rating}/5",
        "",
        f"Original query:",
        query_text or "(not provided)",
        "",
        f"What worked:",
        what_worked or "(blank)",
        "",
        f"What didn't:",
        what_didnt or "(blank)",
    ]
    if xlsx_path:
        body_lines.extend(["", f"Deliverable: {xlsx_path}"])
    body = "\n".join(body_lines)

    fallback_md_lines = [
        f"# internal-comps feedback — {today}",
        "",
        f"- **Rating:** {rating}/5",
        f"- **Deliverable:** {xlsx_path or 'n/a'}",
        "",
        "## Original query",
        "",
        f"> {query_text}" if query_text else "(not provided)",
        "",
        "## What worked",
        "",
        what_worked or "(blank)",
        "",
        "## What didn't",
        "",
        what_didnt or "(blank)",
    ]
    fallback_content = "\n".join(fallback_md_lines)

    fallback_filename = f"feedback-{today}.md"
    fallback_dir = os.path.dirname(xlsx_path) if xlsx_path else os.getcwd()

    return {
        "recipient": "david@groundedintelligence.io",
        "subject": subject,
        "body": body,
        "fallback_filename": fallback_filename,
        "fallback_content": fallback_content,
        "fallback_dir": fallback_dir,
    }
