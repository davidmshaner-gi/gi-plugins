"""External-comps zero-placeholder exclusion (gi-plugins#82).

External-comps computes its broker-facing summary stats inline: the Summary
sheet in format_excel and the "Quick read" range in markdown_table. A 0 rate or
0 SF is an "unknown-value" placeholder, not a real value — including it printed
"Median price_per_sf: 0" on the workbook and collapsed the quick-read range to
"$0.00–...". These tests pin positive-only extraction in both surfaces.

Card: davidmshaner-gi/gi-plugins#82.
"""

import os
import tempfile

from openpyxl import load_workbook

import helpers


def test_quick_read_excludes_zero_rate_placeholder():
    top = [
        {"property_address": "1 A St", "price_per_sf": 100},
        {"property_address": "2 B St", "price_per_sf": 300},
        {"property_address": "3 C St", "price_per_sf": 0},  # placeholder
    ]
    validated = {"transaction_type": "sale"}
    md = helpers.markdown_table(top, [], [], [], validated)
    assert "Quick read" in md
    # Range is over positives (100, 300) -> must NOT start at $0.00.
    assert "$0.00" not in md.split("Quick read")[1]
    assert "$100.00" in md and "$300.00" in md


def test_summary_sheet_excludes_zero_placeholders():
    rows = [
        {"price_per_sf": 100, "building_sf": 10_000},
        {"price_per_sf": 300, "building_sf": 30_000},
        {"price_per_sf": 0, "building_sf": 0},  # placeholder
    ]
    validated = {
        "transaction_type": "sale",
        "asset_type": "Industrial",
        "geography": {"cities": ["Raleigh"]},
        "size_range": {},
        "date_window": {},
    }
    # format_excel's Windows long-path guard (gi-plugins#7) flattens the output to a
    # basename in the CWD, so run inside the tempdir to keep the artifact contained.
    cwd = os.getcwd()
    with tempfile.TemporaryDirectory() as d:
        os.chdir(d)
        try:
            out = helpers.format_excel(rows, validated, "c.xlsx", [], [], [])
            wb = load_workbook(out)
        finally:
            os.chdir(cwd)
        summary = wb["Summary"]
        # Map "Label" -> value from the Summary sheet's two-column layout.
        vals = {}
        for row in summary.iter_rows(min_row=1, max_col=2, values_only=True):
            if row[0] is not None:
                vals[row[0]] = row[1]

    assert vals["Count"] == 3  # row count unchanged
    # Median price_per_sf over positives (100, 300) -> 200, NOT 0.
    assert vals["Median price_per_sf"] == 200
    assert vals["Avg price_per_sf"] == 200
    assert vals["Min price_per_sf"] == 100
    assert vals["Median building_sf"] == 20_000
