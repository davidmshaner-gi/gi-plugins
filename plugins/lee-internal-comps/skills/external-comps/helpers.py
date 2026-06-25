"""
helpers.py — atomic helpers for the external-comps skill.

Run in the Cowork sandbox. The model orchestrates; helpers are deterministic.
None of these helpers call MCP tools — MCP invocation and email sending are
the model's responsibility (the model has MCP access; the sandbox does not).

Design contract:
  - Open-shaped dicts in / dicts out. Helpers tolerate extra keys.
  - Three load-bearing keys on the request: asset_type, transaction_type,
    geography. Everything else is optional and may be absent.
  - Frozen output layout — see SKILL.md "Output" section. Do not parameterize
    beyond what these signatures expose.
"""

from __future__ import annotations

import base64
import json
import os
import statistics
import tempfile
from datetime import date, timedelta
from typing import Any, Iterable, Literal, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.drawing.image import Image as XLImage


# See the canonical XLSX_STUB note in internal-comps/helpers.py. Cowork's per-session
# output dir on Windows runs ~190-210 chars deep (Bonner, 2026-06-25) and Excel refuses
# to OPEN a workbook whose full path exceeds 218 chars; the only lever we control is the
# filename, so we emit the shortest practical stub `c.xlsx`, not a descriptive name, to
# clear 218 even on the deepest session dirs (gi-plugins#7).
XLSX_STUB = "c"  # keep <=2 chars: budget under 218 is ~8 on the deepest broker session dirs


def safe_xlsx_name(path: str = "") -> str:
    """Return the shortest stable .xlsx filename in the CWD, enumerating on collision.

    Emits `c.xlsx` (XLSX_STUB + ".xlsx"); enumerates `c1.xlsx`, `c2.xlsx`, ... if that
    name is already taken in the CWD (a second comps pull in the same session), so a
    later pull never clobbers an earlier deliverable. `path` is accepted for back-compat and
    deliberately ignored — the descriptive name can't survive the Windows 218-char
    Excel-open limit (gi-plugins#7).

    Canonical definition lives in the `internal-comps` skill's helpers.py; this is a
    local mirror (kept identical) because `external-comps` is a standalone broker skill
    and must not take a runtime import dependency on a sibling skill. Same duplication
    idiom as LEE_BRAND_MAROON / the logo asset.
    """
    candidate = f"{XLSX_STUB}.xlsx"
    n = 1
    while os.path.exists(candidate):
        candidate = f"{XLSX_STUB}{n}.xlsx"
        n += 1
    return candidate


LEE_BRAND_MAROON = "98002E"  # official Lee Red, PMS 202 (lee-and-associates#28 / Brand Guidelines)
LEE_LOGO_FILENAME = "lee_logo.png"

# RDU MSA county whitelist — applied post-fetch when geography.named_market resolves to RDU.
RDU_MSA_COUNTIES = {"Wake", "Durham", "Orange", "Chatham", "Johnston", "Franklin", "Granville"}

# Aliases for the RDU named market — case-insensitive match in validate_request.
RDU_MARKET_ALIASES = {"rdu msa", "rdu", "triangle", "raleigh-durham", "raleigh durham"}

# RDU MSA city → county fallback map. Surfaced in SKILL.md so brokers can
# read/override it. Passed to apply_post_filters via the city_to_county
# parameter when the skill needs to enrich null county values.
RDU_CITY_TO_COUNTY: dict[str, str] = {
    # Wake
    "Raleigh": "Wake", "Cary": "Wake", "Garner": "Wake", "Apex": "Wake",
    "Wake Forest": "Wake", "Holly Springs": "Wake", "Morrisville": "Wake",
    "Knightdale": "Wake", "Rolesville": "Wake", "Wendell": "Wake",
    "Zebulon": "Wake", "Fuquay-Varina": "Wake", "Fuquay Varina": "Wake",
    # Durham
    "Durham": "Durham",
    # Orange
    "Chapel Hill": "Orange", "Carrboro": "Orange", "Hillsborough": "Orange",
    # Chatham
    "Pittsboro": "Chatham", "Siler City": "Chatham",
    # Johnston
    "Smithfield": "Johnston", "Clayton": "Johnston", "Selma": "Johnston",
    "Benson": "Johnston", "Four Oaks": "Johnston",
    # Franklin
    "Louisburg": "Franklin", "Youngsville": "Franklin", "Bunn": "Franklin",
    # Granville
    "Creedmoor": "Granville", "Oxford": "Granville", "Butner": "Granville",
}

# When the share of rows with null county exceeds this threshold AND a county
# filter is requested, the model surfaces a 3-strategy dialog before applying
# the filter (see SKILL.md Process step 7). Below threshold, silent enrichment.
NULL_COUNTY_DIALOG_THRESHOLD = 0.20

# CoStar property_type taxonomy — values are the verbatim strings the MCP expects.
# Maps the skill's `asset_type` (broker shorthand) to CoStar's `property_type`.
ASSET_TYPE_TO_COSTAR_SALE = {
    "industrial": "Industrial",
    "office": "Office",
    "retail": "Retail",
    "flex": "Flex",
    "multifamily": "Multifamily",
    "student": "Student",
    "land": "Land",
    "hospitality": "Hospitality",
    "health_care": "Health Care",
    "specialty": "Specialty",
}

ASSET_TYPE_TO_COSTAR_LEASE = {
    "industrial": "Industrial",
    "office": "Office",
    "retail": "Retail",
    "flex": "Flex",
    "medical": "Medical",
}

# Default lookback windows in months.
DEFAULT_SALE_LOOKBACK_MONTHS = 12
DEFAULT_LEASE_LOOKBACK_MONTHS = 6
DEFAULT_TARGET_COUNT = 8
DEFAULT_MIN_SALE_PRICE = 500_000

# Display column order for the main Excel sheet.
DISPLAY_COLUMNS_SALE = [
    "external_id", "property_address", "property_city", "county", "submarket",
    "property_type", "property_secondary_type", "building_sf", "year_built",
    "sale_price", "price_per_sf", "sale_date", "actual_cap_rate", "noi",
    "percent_leased", "sale_type", "sale_conditions", "days_on_market",
    "buyer_true_company", "seller_true_company", "listing_broker_company",
    "costar_property_url", "sale_notes",
]

DISPLAY_COLUMNS_LEASE = [
    "external_id", "property_address", "property_city", "county", "submarket",
    "property_type", "building_sf", "lease_start_date", "lease_term_months",
    "lease_expiration_date", "base_rent", "rent_type", "escalations",
    "free_rent_months", "ti_allowance", "tenant_name", "tenant_industry",
    "floor", "suite", "space_type", "costar_property_url",
]


def validate_request(parsed: dict) -> dict:
    """Apply defaults, list missing required keys and warnings.

    Returns a dict shaped:
      {
        "validated": {...},          # the request with defaults applied
        "missing_required": [str],   # blocking — model should clarify
        "applied_defaults": [str],   # surfaced in the email body
        "warnings": [str],           # non-blocking
      }
    """
    validated = dict(parsed)  # shallow copy; preserve broker input
    missing: list[str] = []
    applied: list[str] = []
    warnings: list[str] = []

    # --- Required keys ---
    asset_type = validated.get("asset_type")
    tx_type = validated.get("transaction_type")
    if not asset_type:
        missing.append("asset_type (industrial, office, retail, flex, medical, etc.)")
    if tx_type not in ("sale", "lease"):
        missing.append("transaction_type (sale or lease)")

    if missing:
        return {
            "validated": validated,
            "missing_required": missing,
            "applied_defaults": applied,
            "warnings": warnings,
        }

    # --- Property-type taxonomy check (warn, don't block) ---
    table = ASSET_TYPE_TO_COSTAR_SALE if tx_type == "sale" else ASSET_TYPE_TO_COSTAR_LEASE
    if asset_type not in table:
        warnings.append(
            f"asset_type '{asset_type}' not in {tx_type} taxonomy "
            f"({sorted(table.keys())}) — passing through verbatim"
        )

    # --- Geography default ---
    if not validated.get("geography"):
        validated["geography"] = {"named_market": "RDU MSA"}
        applied.append("geography → RDU MSA")
    else:
        geo = validated["geography"]
        if "named_market" in geo and geo["named_market"].strip().lower() in RDU_MARKET_ALIASES:
            geo["named_market"] = "RDU MSA"  # normalize

    # --- Date window default ---
    if not validated.get("date_window"):
        months = DEFAULT_SALE_LOOKBACK_MONTHS if tx_type == "sale" else DEFAULT_LEASE_LOOKBACK_MONTHS
        validated["date_window"] = {"lookback_months": months}
        applied.append(f"date_window → past {months} months")

    # --- Target count default ---
    if "target_count" not in validated:
        validated["target_count"] = DEFAULT_TARGET_COUNT
        applied.append(f"target_count → {DEFAULT_TARGET_COUNT}")

    # --- Sale-only min price default (junk filter) ---
    if tx_type == "sale" and "min_sale_price" not in validated:
        validated["min_sale_price"] = DEFAULT_MIN_SALE_PRICE
        applied.append(f"min_sale_price → ${DEFAULT_MIN_SALE_PRICE:,} (junk filter)")

    # --- Size warning ---
    if not validated.get("size_range"):
        warnings.append("size_range not specified — results may be wide; ask if needed")

    return {
        "validated": validated,
        "missing_required": [],
        "applied_defaults": applied,
        "warnings": warnings,
    }


def _resolve_date_window(window: dict, tx_type: str = "sale", anchor: Optional[date] = None) -> tuple[str, str]:
    """Resolve a date_window dict to (from_iso, to_iso) ISO date strings."""
    anchor = anchor or date.today()
    if "from" in window and "to" in window:
        return window["from"], window["to"]
    default_months = DEFAULT_SALE_LOOKBACK_MONTHS if tx_type == "sale" else DEFAULT_LEASE_LOOKBACK_MONTHS
    months = int(window.get("lookback_months", default_months))
    # Approximate month math — good enough for comp queries.
    days = months * 30
    from_d = anchor - timedelta(days=days)
    return from_d.isoformat(), anchor.isoformat()


def build_mcp_params(validated: dict) -> dict:
    """Map a validated request to (tool_name, params_dict, city_list).

    Returns:
      {
        "tool_name": "search_external_sale_comps" | "search_external_lease_comps",
        "params_list": [params_dict, ...],   # one entry per MCP call needed
                                             # (1 call for named_market; N calls for cities)
        "post_filter_counties": set[str] | None,  # apply post-fetch if named_market
      }

    The model loops `params_list`, invoking the MCP tool for each entry, then unions
    the rows before calling apply_post_filters.
    """
    tx = validated["transaction_type"]
    asset = validated["asset_type"]
    tool_name = (
        "search_external_sale_comps" if tx == "sale" else "search_external_lease_comps"
    )

    # --- Property type ---
    table = ASSET_TYPE_TO_COSTAR_SALE if tx == "sale" else ASSET_TYPE_TO_COSTAR_LEASE
    property_type = table.get(asset, asset.title())  # fallback: title-case the broker value

    # --- Date window ---
    date_from, date_to = _resolve_date_window(validated["date_window"], tx_type=tx)

    # --- Build common params (everything except city) ---
    base: dict[str, Any] = {"property_type": property_type, "state": "NC"}

    if tx == "sale":
        base["min_sale_date"] = date_from
        base["max_sale_date"] = date_to
    else:
        base["min_lease_start_date"] = date_from
        base["max_lease_start_date"] = date_to

    if validated.get("size_range"):
        sr = validated["size_range"]
        if "min_sf" in sr:
            base["min_building_sf"] = int(sr["min_sf"])
        if "max_sf" in sr:
            base["max_building_sf"] = int(sr["max_sf"])

    # --- Transaction-type specific filters ---
    if tx == "sale":
        if "min_sale_price" in validated:
            base["min_sale_price"] = int(validated["min_sale_price"])
        if "max_sale_price" in validated:
            base["max_sale_price"] = int(validated["max_sale_price"])
        if "min_cap_rate" in validated:
            base["min_cap_rate"] = float(validated["min_cap_rate"])
        if "max_cap_rate" in validated:
            base["max_cap_rate"] = float(validated["max_cap_rate"])
    else:
        if "min_base_rent" in validated:
            base["min_base_rent"] = float(validated["min_base_rent"])
        if "max_base_rent" in validated:
            base["max_base_rent"] = float(validated["max_base_rent"])
        if "min_lease_term_months" in validated:
            base["min_lease_term_months"] = int(validated["min_lease_term_months"])
        if "max_lease_term_months" in validated:
            base["max_lease_term_months"] = int(validated["max_lease_term_months"])
        if "tenant_industry" in validated:
            base["tenant_industry"] = validated["tenant_industry"]

    base["limit"] = 200  # always pull the max; rank trims to the sweet spot

    # --- Geography → params_list + post-filter set ---
    geo = validated["geography"]
    params_list: list[dict] = []
    post_filter_counties: Optional[set[str]] = None

    if "named_market" in geo:
        # Named market → one call, no city, post-filter by county whitelist.
        params_list.append(dict(base))
        if geo["named_market"] == "RDU MSA":
            post_filter_counties = set(RDU_MSA_COUNTIES)
    elif "cities" in geo and geo["cities"]:
        # Cities → one call per city, no county filter post-fetch.
        for city in geo["cities"]:
            p = dict(base)
            p["city"] = city
            params_list.append(p)
    else:
        # Geography present but empty — fall through to no-geo (state-only).
        params_list.append(dict(base))

    return {
        "tool_name": tool_name,
        "params_list": params_list,
        "post_filter_counties": post_filter_counties,
    }


def null_county_rate(rows: list[dict]) -> tuple[int, int, float]:
    """Return (null_count, total_count, share). `share` is null_count / total_count
    (0.0 when total_count == 0). `county` is considered null if absent, None, or
    blank (whitespace-only). The model uses `share` to decide whether to surface
    the 3-strategy null-county dialog (see SKILL.md Process step 7).
    """
    total = len(rows)
    if total == 0:
        return 0, 0, 0.0
    null_count = sum(1 for r in rows if not (r.get("county") or "").strip())
    return null_count, total, null_count / total


def apply_post_filters(
    rows: list[dict],
    validated: dict,
    post_filter_counties: Optional[set[str]] = None,
    city_to_county: Optional[dict[str, str]] = None,
) -> tuple[list[dict], list[str]]:
    """Apply Python-side filters that the MCP couldn't express.

    Currently: county whitelist (for named_market="RDU MSA"), with optional
    city → county fallback when the snapshot has null counties.

    Args:
      rows: MCP-returned rows.
      validated: the validated request dict (currently unused but reserved for
        future filter axes like sub-region exclusions).
      post_filter_counties: county whitelist. None = skip the county filter.
      city_to_county: optional fallback map. When set, rows with null/empty
        `county` are enriched in-place from `property_city` before filtering.
        Rows whose city isn't in the map keep null county and fall through to
        the existing "not in whitelist → drop" behavior. Note: if zero rows
        are enriched (no nulls, or all unmapped), no "inferred" entry is added
        to `applied_filters` — the drop count alone tells the story.

    Returns (filtered_rows, applied_filters_log) where applied_filters_log is a
    list of human-readable strings describing what was filtered, for the email
    body and Methodology sheet.
    """
    applied: list[str] = []
    out = rows

    if post_filter_counties:
        # --- Optional pre-pass: enrich null counties from city map ---
        inferred = 0
        if city_to_county:
            for r in out:
                if not (r.get("county") or "").strip():
                    city = (r.get("property_city") or "").strip()
                    if city and city in city_to_county:
                        r["county"] = city_to_county[city]
                        inferred += 1
            if inferred > 0:
                applied.append(
                    f"inferred {inferred} county value{'s' if inferred != 1 else ''} "
                    f"from property_city (RDU map)"
                )

        # --- County whitelist filter ---
        before = len(out)
        out = [r for r in out if (r.get("county") or "").strip() in post_filter_counties]
        dropped = before - len(out)
        if dropped > 0:
            applied.append(
                f"dropped {dropped} row{'s' if dropped != 1 else ''} outside {sorted(post_filter_counties)}"
            )

    return out, applied


def _months_since(date_str: Optional[str], anchor: Optional[date] = None) -> float:
    """Months between an ISO date string and anchor. Returns +inf if unparsable."""
    if not date_str:
        return float("inf")
    try:
        d = date.fromisoformat(date_str[:10])
    except (ValueError, TypeError):
        return float("inf")
    anchor = anchor or date.today()
    return (anchor - d).days / 30.0


def rank_comps(
    rows: list[dict],
    validated: dict,
) -> tuple[list[dict], list[dict], list[dict], list[dict]]:
    """Split rows into ranked-top, tagged-under-contract, tagged-sublet,
    tagged-rent-undisclosed. Returns four lists.

    Tagging rules (borrowed from stashed costar-comps flows):
      - Sale: rows with sale_conditions LIKE 'Under Contract%' → tagged_under_contract.
      - Lease: rows with rent_type == 'Sublet' → tagged_sublet.
      - Lease: rows where base_rent is missing → tagged_rent_undisclosed.

    Composite score (ascending):
      recency * 0.4 + size_proximity * 0.3 + geo * 0.3
    """
    tx = validated["transaction_type"]
    size_range = validated.get("size_range") or {}
    target_size = None
    if "min_sf" in size_range and "max_sf" in size_range:
        target_size = (int(size_range["min_sf"]) + int(size_range["max_sf"])) / 2.0

    tagged_uc: list[dict] = []
    tagged_sublet: list[dict] = []
    tagged_rent_undisclosed: list[dict] = []
    main: list[dict] = []

    for r in rows:
        if tx == "sale":
            cond = (r.get("sale_conditions") or "").lower()
            if "under contract" in cond:
                tagged_uc.append(r)
                continue
        else:
            rent_type = (r.get("rent_type") or "").lower()
            if rent_type == "sublet":
                tagged_sublet.append(r)
                continue
            if r.get("base_rent") in (None, "", 0):
                tagged_rent_undisclosed.append(r)
                continue
        main.append(r)

    # --- Score the main set ---
    date_key = "sale_date" if tx == "sale" else "lease_start_date"

    def score(r: dict) -> float:
        recency = _months_since(r.get(date_key))
        if target_size and r.get("building_sf"):
            size_prox = abs(float(r["building_sf"]) - target_size)
        else:
            size_prox = 0.0  # no target → skip the term
        # Geo score: core RDU submarkets get 0; edge gets 1.
        submarket = (r.get("submarket") or "").lower()
        core_submarkets = {"glenwood", "creedmoor", "south durham", "rtp",
                           "se wake", "north hills", "downtown raleigh"}
        geo = 0.0 if any(k in submarket for k in core_submarkets) else 1.0
        # v1 normalization: size_prox / 1000. Practically fine because MCP size filters bound the spread; revisit if rankings feel off.
        return recency * 0.4 + (size_prox / 1000.0) * 0.3 + geo * 0.3

    main_sorted = sorted(main, key=score)
    return main_sorted, tagged_uc, tagged_sublet, tagged_rent_undisclosed


def _decode_logo_to_temp() -> Optional[str]:
    """Write the bundled logo to a temp file and return its path.

    The sibling internal-comps helpers.py embeds the logo as base64; we read the
    file directly from the skill bundle here (Cowork makes skill files available).
    Falls back to None if the logo isn't present — Excel still works, just no logo.
    """
    logo_path = os.path.join(os.path.dirname(__file__), LEE_LOGO_FILENAME)
    if os.path.isfile(logo_path):
        return logo_path
    return None


def _sheet_title(validated: dict) -> str:
    asset = validated["asset_type"].replace("_", " ").title()
    tx = "Sale" if validated["transaction_type"] == "sale" else "Lease"
    geo = validated["geography"]
    if "named_market" in geo:
        geo_str = geo["named_market"]
    elif "cities" in geo:
        geo_str = ", ".join(geo["cities"])
    else:
        geo_str = "NC"
    title = f"{asset} {tx} {geo_str} Comps"
    return title[:31]  # Excel sheet name limit


def format_excel(
    rows: list[dict],
    validated: dict,
    xlsx_path: str,
    applied_defaults: list[str],
    warnings: list[str],
    applied_filters: list[str],
    last_sync: Optional[str] = None,
) -> str:
    """Write the 3-sheet workbook. Returns xlsx_path.

    Layout is frozen — matches internal-comps' visual language (Lee-branded
    header in official Lee Red with the logo + title band on the main sheet,
    frozen panes, autofilter, color scale on rate column).
    """
    tx = validated["transaction_type"]
    cols = DISPLAY_COLUMNS_SALE if tx == "sale" else DISPLAY_COLUMNS_LEASE
    rate_col = "price_per_sf" if tx == "sale" else "base_rent"

    wb = Workbook()
    ws = wb.active
    ws.title = _sheet_title(validated)

    # --- Lee branding: logo + title band at top, then header row, then data ---
    # Mirrors internal-comps (gi-plugins#90). Official Lee Red #98002E (PMS 202).
    logo_path = _decode_logo_to_temp()
    logo_available = logo_path is not None
    if not logo_available:
        # Non-fatal: the workbook still generates, but a silent unbranded
        # workbook must never ship unnoticed again (gi-plugins#90).
        warnings.append("Lee logo asset unavailable — workbook generated without branding.")
    header_row_idx = 4 if logo_available else 1

    if logo_available:
        ws.row_dimensions[1].height = 56
        try:
            img = XLImage(logo_path)
            img.width = 180
            img.height = 60
            ws.add_image(img, "A1")
        except Exception:
            pass  # logo is decoration; never block the workbook
        title_kind = "Sale Comps" if tx == "sale" else "Lease Comps"
        ws.cell(row=2, column=2, value=_sheet_title(validated)).font = Font(
            bold=True, size=14, color=LEE_BRAND_MAROON
        )
        ws.cell(row=3, column=2, value=f"Pulled {date.today().isoformat()} · External {title_kind}").font = Font(
            italic=True, size=10, color="555555"
        )

    # --- Header row ---
    header_fill = PatternFill("solid", fgColor=LEE_BRAND_MAROON)
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col_idx, col_name in enumerate(cols, start=1):
        cell = ws.cell(row=header_row_idx, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # --- Data rows ---
    data_start = header_row_idx + 1
    for row_offset, r in enumerate(rows):
        excel_row = data_start + row_offset
        for col_idx, col_name in enumerate(cols, start=1):
            ws.cell(row=excel_row, column=col_idx, value=r.get(col_name))

    # --- Freeze panes + autofilter (relative to the header row) ---
    ws.freeze_panes = f"A{data_start}"
    last_row = data_start + len(rows) - 1 if rows else header_row_idx
    ws.auto_filter.ref = f"A{header_row_idx}:{get_column_letter(len(cols))}{last_row}"

    # --- Color scale on rate column ---
    if rate_col in cols and len(rows) > 0:
        rate_col_letter = get_column_letter(cols.index(rate_col) + 1)
        rule_range = f"{rate_col_letter}{data_start}:{rate_col_letter}{last_row}"
        ws.conditional_formatting.add(
            rule_range,
            ColorScaleRule(
                start_type="min", start_color="F8696B",
                mid_type="percentile", mid_value=50, mid_color="FFEB84",
                end_type="max", end_color="63BE7B",
            ),
        )

    # --- Column widths ---
    for col_idx, col_name in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, len(col_name) + 2)

    # --- Sheet 2: Summary ---
    summary = wb.create_sheet("Summary")
    summary.cell(row=1, column=1, value="Count").font = Font(bold=True)
    summary.cell(row=1, column=2, value=len(rows))

    # gi-plugins#82: comps metrics are physically positive — a 0 rate or 0 SF is an
    # "unknown-value" placeholder, never a real value. Excluding them from the Summary
    # stats keeps "Median $/SF: $0.00" off broker-facing workbooks. Count stays len(rows).
    rate_vals = [r.get(rate_col) for r in rows
                 if isinstance(r.get(rate_col), (int, float)) and r.get(rate_col) > 0]
    size_vals = [r.get("building_sf") for r in rows
                 if isinstance(r.get("building_sf"), (int, float)) and r.get("building_sf") > 0]

    def _stat(label: str, fn, vals, row_idx: int) -> None:
        summary.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        summary.cell(row=row_idx, column=2, value=(fn(vals) if vals else None))

    _stat(f"Avg {rate_col}", statistics.mean, rate_vals, 2)
    _stat(f"Median {rate_col}", statistics.median, rate_vals, 3)
    _stat(f"Min {rate_col}", min, rate_vals, 4)
    _stat(f"Max {rate_col}", max, rate_vals, 5)
    _stat("Avg building_sf", statistics.mean, size_vals, 6)
    _stat("Median building_sf", statistics.median, size_vals, 7)

    # --- Sheet 3: Methodology ---
    meth = wb.create_sheet("Methodology")
    pairs = [
        ("Pulled for", "Lee & Associates Raleigh broker"),
        ("Pull date", date.today().isoformat()),
        ("Source", "External weekly snapshot via lee-raleigh-mcp"),
        ("Geography", json.dumps(validated.get("geography"))),
        ("Property type", validated.get("asset_type")),
        ("Size range", json.dumps(validated.get("size_range"))),
        ("Date window", json.dumps(validated.get("date_window"))),
        ("Applied defaults", "; ".join(applied_defaults) or "—"),
        ("Warnings", "; ".join(warnings) or "—"),
        ("Applied filters", "; ".join(applied_filters) or "—"),
        ("Last snapshot sync", last_sync or "unknown"),
        ("Caveat", "External comps data ingested weekly. Latest snapshot only."),
    ]
    for i, (k, v) in enumerate(pairs, start=1):
        meth.cell(row=i, column=1, value=k).font = Font(bold=True)
        meth.cell(row=i, column=2, value=v)
    meth.column_dimensions["A"].width = 22
    meth.column_dimensions["B"].width = 70

    # Windows 218-char path guard (shared helper). Flatten to a CWD basename and
    # cap the filename so a deep/long path can't survive (gi-plugins#7).
    xlsx_path = safe_xlsx_name(xlsx_path)
    wb.save(xlsx_path)
    return xlsx_path


def _fmt_money(v: Any) -> str:
    if isinstance(v, (int, float)):
        return f"${v:,.0f}"
    return "—"


def _fmt_rate(v: Any) -> str:
    if isinstance(v, (int, float)):
        return f"${v:,.2f}"
    return "—"


def _fmt_int(v: Any) -> str:
    if isinstance(v, (int, float)):
        return f"{int(v):,}"
    return "—"


def markdown_table(
    top: list[dict],
    tagged_uc: list[dict],
    tagged_sublet: list[dict],
    tagged_rent_undisclosed: list[dict],
    validated: dict,
) -> str:
    """Build the Markdown reply: main ranked table plus tagged sub-tables plus quick read."""
    tx = validated["transaction_type"]
    target_count = validated.get("target_count", DEFAULT_TARGET_COUNT)
    top_n = top[:target_count]

    parts: list[str] = []

    # --- Main table ---
    if tx == "sale":
        parts.append("| # | Address | City | County | SF | Sale Price | $/SF | Sale Date | YB | Type | Submarket | Notes |")
        parts.append("|---|---|---|---|---|---|---|---|---|---|---|---|")
        for i, r in enumerate(top_n, start=1):
            parts.append(
                f"| {i} | {r.get('property_address','—')} | {r.get('property_city','—')} | "
                f"{r.get('county','—')} | {_fmt_int(r.get('building_sf'))} | "
                f"{_fmt_money(r.get('sale_price'))} | {_fmt_rate(r.get('price_per_sf'))} | "
                f"{r.get('sale_date','—')} | {r.get('year_built','—')} | "
                f"{r.get('property_type','—')} | {r.get('submarket','—')} | "
                f"{(r.get('sale_notes') or '')[:60]} |"
            )
    else:
        parts.append("| # | Address | City | County | SF | Rent ($/SF) | Type | Term (mo) | Tenant | Signed | Notes |")
        parts.append("|---|---|---|---|---|---|---|---|---|---|---|")
        for i, r in enumerate(top_n, start=1):
            parts.append(
                f"| {i} | {r.get('property_address','—')} | {r.get('property_city','—')} | "
                f"{r.get('county','—')} | {_fmt_int(r.get('building_sf'))} | "
                f"{_fmt_rate(r.get('base_rent'))} | {r.get('rent_type','—')} | "
                f"{r.get('lease_term_months','—')} | {r.get('tenant_name','—')} | "
                f"{r.get('lease_start_date','—')} | {(r.get('tenant_industry') or '')[:40]} |"
            )

    # --- Tagged sub-tables ---
    if tagged_uc:
        parts.append("\n**Under Contract** (not yet closed)")
        parts.append(f"_{len(tagged_uc)} rows — see Excel for full details._")
    if tagged_sublet:
        parts.append("\n**Sublet** (not direct deals)")
        parts.append(f"_{len(tagged_sublet)} rows — see Excel for full details._")
    if tagged_rent_undisclosed:
        parts.append("\n**Rent not disclosed** (activity signal only)")
        parts.append(f"_{len(tagged_rent_undisclosed)} rows — see Excel for full details._")

    # --- Quick read ---
    rate_key = "price_per_sf" if tx == "sale" else "base_rent"
    # gi-plugins#82: a 0 rate is an unknown-value placeholder; excluding it keeps the
    # quick-read range from collapsing to "$0.00–..." on the broker-facing email.
    rates = [r.get(rate_key) for r in top_n
             if isinstance(r.get(rate_key), (int, float)) and r.get(rate_key) > 0]
    if rates:
        lo, hi = min(rates), max(rates)
        parts.append(f"\n**Quick read:** {rate_key} ranges {_fmt_rate(lo)}–{_fmt_rate(hi)} across {len(top_n)} comps.")

    return "\n".join(parts)


def draft_email(
    filtered_rows: list[dict],
    top: list[dict],
    validated: dict,
    xlsx_path: Optional[str],
    applied_defaults: list[str],
    warnings: list[str],
    applied_filters: list[str],
) -> dict:
    """Returns {subject, body}. Body surfaces total filtered count, top-N ranked count,
    stats, defaults, warnings, filters, and a widen-question if below target_count."""
    tx = validated["transaction_type"]
    asset = validated["asset_type"]
    geo = validated["geography"]
    geo_str = geo.get("named_market") or ", ".join(geo.get("cities", []) or ["NC"])
    target = validated.get("target_count", DEFAULT_TARGET_COUNT)
    total_count = len(filtered_rows)
    top_count = len(top)

    subject = f"External {tx} comps — {asset} {geo_str} ({total_count})"

    body_lines = [
        f"Pulled {total_count} external {tx} comp{'s' if total_count != 1 else ''} "
        f"for {asset} in {geo_str} (external weekly snapshot via lee-raleigh-mcp). "
        f"Top {top_count} ranked in the attached table.",
    ]

    if applied_defaults:
        body_lines.append("")
        body_lines.append("Defaults applied (push back any time):")
        for d in applied_defaults:
            body_lines.append(f"  - {d}")

    if warnings:
        body_lines.append("")
        body_lines.append("Notes:")
        for w in warnings:
            body_lines.append(f"  - {w}")

    if applied_filters:
        body_lines.append("")
        body_lines.append("Filters applied post-fetch:")
        for f in applied_filters:
            body_lines.append(f"  - {f}")

    if top_count < target:
        body_lines.append("")
        body_lines.append(
            f"Below target ({top_count} vs. {target}). Want me to widen on size, "
            f"date, or geography? Let me know which."
        )

    if xlsx_path:
        body_lines.append("")
        body_lines.append(f"Excel attached: {os.path.basename(xlsx_path)}")

    if validated.get("notes"):
        body_lines.append("")
        body_lines.append(f"Your notes back: {validated['notes']}")

    return {"subject": subject, "body": "\n".join(body_lines)}


def format_feedback(
    request_summary: str,
    rating: Optional[int],
    what_worked: Optional[str],
    what_didnt: Optional[str],
    broker_email: Optional[str] = None,
) -> dict:
    """Returns a structured feedback payload.

    Shape:
      {
        "email_to": "david@groundedintelligence.io",
        "email_subject": "...",
        "email_body": "...",
        "fallback_filename": "feedback-YYYY-MM-DD.md",
        "fallback_content": "..."
      }
    """
    today = date.today().isoformat()
    subject = f"[external-comps feedback] {request_summary[:60]}"
    body_parts = [
        f"Date: {today}",
        f"Broker: {broker_email or 'not captured'}",
        f"Request: {request_summary}",
        f"Rating (1-5): {rating if rating is not None else 'skipped'}",
        f"What worked: {what_worked or 'skipped'}",
        f"What didn't: {what_didnt or 'skipped'}",
    ]
    body = "\n".join(body_parts)
    return {
        "email_to": "david@groundedintelligence.io",
        "email_subject": subject,
        "email_body": body,
        "fallback_filename": f"feedback-{today}.md",
        "fallback_content": body,
    }
